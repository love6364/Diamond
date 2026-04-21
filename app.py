
# import streamlit as st
# import pandas as pd

# st.title("💎 Diamond Full Automation System")

# # Upload files
# main_file = st.file_uploader("Upload Diamond File", type=["xlsx"])
# st_file = st.file_uploader("Upload ST File (Status)", type=["xlsx"])
# days_file = st.file_uploader("Upload Days File", type=["xlsx"])

# # Size groups
# size_groups = [
#     (0.30, 0.39),(0.40, 0.49),(0.50, 0.59),(0.60, 0.69),
#     (0.70, 0.79),(0.80, 0.89),(0.90, 0.99),
#     (1.00, 1.10),(1.11, 1.49),(1.50, 1.59),
#     (1.60, 1.69),(1.70, 1.79),(1.80, 1.89),
#     (1.90, 1.99),(2.00, 2.49),(2.50, 2.99),
#     (3.00, 3.99),(4.00, 5.00)
# ]

# def assign_group(carat):
#     for low, high in size_groups:
#         if low <= carat <= high:
#             return f"{low:.2f}-{high:.2f}"
#     return "Other"

# if main_file and st_file and days_file:

#     # Read files
#     df = pd.read_excel(main_file)
#     st_df = pd.read_excel(st_file)
#     days_df = pd.read_excel(days_file)

#     st.write("### 📊 Main Data")
#     st.dataframe(df)

#     # Column selection
#     lot_main = st.selectbox("Lot No (Main File)", df.columns)
#     lot_st = st.selectbox("Lot No (ST File)", st_df.columns)
#     status_col = st.selectbox("Status Column (ST)", st_df.columns)
#     lot_days = st.selectbox("Lot No (Days File)", days_df.columns)
#     days_col = st.selectbox("Days Column", days_df.columns)
#     carat_col = st.selectbox("Carat Column", df.columns)

#     # 🔥 CLEAN CARAT
#     df[carat_col] = (
#         df[carat_col]
#         .astype(str)
#         .str.replace("ct", "", regex=False)
#         .str.replace(",", ".")
#     )
#     df[carat_col] = pd.to_numeric(df[carat_col], errors='coerce')

#     # 🔥 STATUS (XLOOKUP)
#     status_map = st_df.set_index(lot_st)[status_col]
#     df["Status"] = df[lot_main].map(status_map)

#     # 🔥 DAYS (XLOOKUP)
#     days_map = days_df.set_index(lot_days)[days_col]
#     df["No_of_Days"] = df[lot_main].map(days_map)

#     # 🔥 MOVE STATUS AFTER LOT
#     lot_index = df.columns.get_loc(lot_main)
#     status_series = df.pop("Status")
#     df.insert(lot_index + 1, "Status", status_series)

#     # 🔥 MOVE DAYS AFTER STATUS
#     status_index = df.columns.get_loc("Status")
#     days_series = df.pop("No_of_Days")
#     df.insert(status_index + 1, "No_of_Days", days_series)

#     # 🔥 SIZE GROUP
#     size_series = df[carat_col].apply(assign_group)

#     # 🔥 INSERT SIZE GROUP AFTER CARAT
#     carat_index = df.columns.get_loc(carat_col)
#     df.insert(carat_index + 1, "Size_Group", size_series)

#     st.write("### ✅ Final Output")
#     st.dataframe(df)

#     # Download
#     output_file = "final_merged_output.xlsx"
#     df.to_excel(output_file, index=False)

#     with open(output_file, "rb") as f:
#         st.download_button("📥 Download Excel", f, output_file)







# import streamlit as st
# import pandas as pd
# import difflib

# # 🔥 PAGE CONFIG
# st.set_page_config(page_title="Diamond Automation", layout="wide")

# # 🔥 UI DESIGN
# st.markdown("""
# <style>
# body {
#     background: linear-gradient(135deg, #0f172a, #1e293b);
# }
# .main-title {
#     font-size: 36px;
#     font-weight: bold;
#     color: #38bdf8;
#     text-align: center;
# }
# .card {
#     background: rgba(255,255,255,0.05);
#     padding: 20px;
#     border-radius: 15px;
#     backdrop-filter: blur(10px);
#     margin-bottom: 20px;
# }
# </style>
# """, unsafe_allow_html=True)

# st.markdown('<div class="main-title">💎 Diamond Automation System</div>', unsafe_allow_html=True)

# # 🔥 FILE UPLOAD
# st.markdown('<div class="card">', unsafe_allow_html=True)
# col1, col2, col3 = st.columns(3)

# with col1:
#     main_file = st.file_uploader("📊 Diamond File", type=["xlsx"])

# with col2:
#     st_file = st.file_uploader("📋 Status File", type=["xlsx"])

# with col3:
#     days_file = st.file_uploader("📅 Days File", type=["xlsx"])

# st.markdown('</div>', unsafe_allow_html=True)

# # 🔥 CLEAN COLUMN NAMES
# def clean_cols(columns):
#     return [col.lower().replace(" ", "").replace("_", "") for col in columns]

# # 🔥 SMART FIND FUNCTION
# def smart_find_column(columns, keywords):
#     clean = clean_cols(columns)
#     for key in keywords:
#         matches = difflib.get_close_matches(key, clean, n=1, cutoff=0.6)
#         if matches:
#             idx = clean.index(matches[0])
#             return columns[idx]
#     return None

# # 🔥 SIZE GROUP FUNCTION
# def assign_group(carat):
#     size_groups = [
#         (0.30,0.39),(0.40,0.49),(0.50,0.59),(0.60,0.69),
#         (0.70,0.79),(0.80,0.89),(0.90,0.99),
#         (1.00,1.10),(1.11,1.49),(1.50,1.99),
#         (2.00,3.00)
#     ]
#     for low, high in size_groups:
#         if low <= carat <= high:
#             return f"{low:.2f}-{high:.2f}"
#     return "Other"

# # 🔥 PROCESS
# if main_file and st_file and days_file:

#     df = pd.read_excel(main_file)
#     st_df = pd.read_excel(st_file)
#     days_df = pd.read_excel(days_file)

#     # 🔥 AUTO DETECT
#     lot_main = smart_find_column(df.columns, ["lot", "lotno", "stoneid"])
#     carat_col = smart_find_column(df.columns, ["carat", "cts", "weight"])

#     lot_st = smart_find_column(st_df.columns, ["lot", "stoneid"])
#     status_col = smart_find_column(st_df.columns, ["status", "st"])

#     lot_days = smart_find_column(days_df.columns, ["lot", "stoneid"])
#     days_col = smart_find_column(days_df.columns, ["No_of_Days"])

#     # 🔥 DEBUG SHOW
#     st.write("🔍 Detected Columns:", {
#         "Lot_Main": lot_main,
#         "Carat": carat_col,
#         "Status": status_col,
#         "Days": days_col
#     })

#     # 🔥 FALLBACK IF NOT FOUND
#     if not lot_main:
#         lot_main = st.selectbox("Select Lot Column (Main)", df.columns)

#     if not carat_col:
#         carat_col = st.selectbox("Select Carat Column", df.columns)

#     if not lot_st:
#         lot_st = st.selectbox("Select Lot Column (ST)", st_df.columns)

#     if not status_col:
#         status_col = st.selectbox("Select Status Column", st_df.columns)

#     if not lot_days:
#         lot_days = st.selectbox("Select Lot Column (Days)", days_df.columns)

#     if not days_col:
#         days_col = st.selectbox("Select Days Column", days_df.columns)

#     # 🔥 BUTTON
#     if st.button("🚀 Generate Final File"):

#         # CLEAN CARAT
#         df[carat_col] = (
#             df[carat_col].astype(str)
#             .str.replace("ct", "", regex=False)
#             .str.replace(",", ".")
#         )
#         df[carat_col] = pd.to_numeric(df[carat_col], errors='coerce')

#         # STATUS (XLOOKUP)
#         status_map = st_df.set_index(lot_st)[status_col]
#         df["Status"] = df[lot_main].map(status_map)

#         # DAYS (XLOOKUP)
#         days_map = days_df.set_index(lot_days)[days_col]
#         df["No_of_Days"] = df[lot_main].map(days_map)

#         # COLUMN POSITION
#         lot_index = df.columns.get_loc(lot_main)
#         df.insert(lot_index + 1, "Status", df.pop("Status"))

#         status_index = df.columns.get_loc("Status")
#         df.insert(status_index + 1, "No_of_Days", df.pop("No_of_Days"))

#         # SIZE GROUP
#         size_series = df[carat_col].apply(assign_group)
#         carat_index = df.columns.get_loc(carat_col)
#         df.insert(carat_index + 1, "Size_Group", size_series)

#         # OUTPUT
#         st.markdown('<div class="card">', unsafe_allow_html=True)
#         st.subheader("📊 Final Output")

#         st.dataframe(df, use_container_width=True)

#         # DOWNLOAD
#         output_file = "final_output.xlsx"
#         df.to_excel(output_file, index=False)

#         with open(output_file, "rb") as f:
#             st.download_button("📥 Download Excel", f, output_file)

#         st.markdown('</div>', unsafe_allow_html=True)



# import streamlit as st
# import pandas as pd
# import difflib

# # 🔥 PAGE CONFIG
# st.set_page_config(page_title="Diamond Excel System", layout="wide")

# # 🔐 LOGIN SYSTEM
# def login():
#     st.markdown("### 🔐 Login to Access System")
#     username = st.text_input("Username")
#     password = st.text_input("Password", type="password")

#     if st.button("Login"):
#         if username == "admin" and password == "1234":
#             st.session_state["login"] = True
#         else:
#             st.error("❌ Invalid Username or Password")

# if "login" not in st.session_state:
#     st.session_state["login"] = False

# if not st.session_state["login"]:
#     login()
#     st.stop()

# # 🔓 LOGOUT
# if st.button("Logout"):
#     st.session_state["login"] = False
#     st.rerun()

# # 🎨 UI DESIGN
# st.markdown("""
# <style>
# body {
#     background: linear-gradient(135deg, #0f172a, #1e293b);
# }
# .title {
#     font-size: 40px;
#     font-weight: bold;
#     text-align: center;
#     color: #38bdf8;
# }
# .subtitle {
#     text-align: center;
#     color: #94a3b8;
#     margin-bottom: 20px;
# }
# .card {
#     background: rgba(255,255,255,0.05);
#     padding: 20px;
#     border-radius: 15px;
#     backdrop-filter: blur(10px);
#     margin-bottom: 20px;
# }
# .upload-box {
#     border: 2px dashed #38bdf8;
#     padding: 20px;
#     border-radius: 12px;
#     text-align: center;
# }
# .stButton>button {
#     background: linear-gradient(90deg, #38bdf8, #6366f1);
#     color: white;
#     border-radius: 10px;
#     height: 45px;
#     font-size: 16px;
#     border: none;
#     width: 100%;
# }
# </style>
# """, unsafe_allow_html=True)

# st.markdown('<div class="title">💎 Diamond\'s Excel Data Generator System</div>', unsafe_allow_html=True)
# st.markdown('<div class="subtitle">Upload → Process → Download</div>', unsafe_allow_html=True)

# # 📂 FILE UPLOAD
# st.markdown('<div class="card">', unsafe_allow_html=True)

# col1, col2, col3 = st.columns(3)

# with col1:
#     st.markdown('<div class="upload-box">📊 Diamond File</div>', unsafe_allow_html=True)
#     main_file = st.file_uploader("", type=["xlsx"], key="main")

# with col2:
#     st.markdown('<div class="upload-box">📋 Status File</div>', unsafe_allow_html=True)
#     st_file = st.file_uploader("", type=["xlsx"], key="st")

# with col3:
#     st.markdown('<div class="upload-box">📅 Days File</div>', unsafe_allow_html=True)
#     days_file = st.file_uploader("", type=["xlsx"], key="days")

# st.markdown('</div>', unsafe_allow_html=True)

# # 🔧 FUNCTIONS
# def clean_cols(columns):
#     return [col.lower().replace(" ", "").replace("_", "") for col in columns]

# def smart_find_column(columns, keywords):
#     clean = clean_cols(columns)
#     for key in keywords:
#         matches = difflib.get_close_matches(key, clean, n=1, cutoff=0.6)
#         if matches:
#             idx = clean.index(matches[0])
#             return columns[idx]
#     return None

# def assign_group(carat):
#     size_groups = [
#         (0.30,0.39),(0.40,0.49),(0.50,0.59),(0.60,0.69),
#         (0.70,0.79),(0.80,0.89),(0.90,0.99),
#         (1.00,1.10),(1.11,1.49),(1.50,1.99),
#         (2.00,3.00)
#     ]
#     for low, high in size_groups:
#         if low <= carat <= high:
#             return f"{low:.2f}-{high:.2f}"
#     return "Other"

# # 🚀 MAIN PROCESS
# if main_file and st_file and days_file:

#     df = pd.read_excel(main_file)
#     st_df = pd.read_excel(st_file)
#     days_df = pd.read_excel(days_file)

#     # 🔍 AUTO DETECT
#     lot_main = smart_find_column(df.columns, ["lot", "lotno", "stoneid"])
#     carat_col = smart_find_column(df.columns, ["carat", "cts", "weight"])

#     lot_st = smart_find_column(st_df.columns, ["lot", "stoneid"])
#     status_col = smart_find_column(st_df.columns, ["status", "st"])

#     lot_days = smart_find_column(days_df.columns, ["lot", "stoneid"])
#     days_col = smart_find_column(days_df.columns, ["day", "days", "no_of_days", "age"])

#     # 🧠 FALLBACK (ONLY IF FAILED)
#     if not lot_main:
#         lot_main = st.selectbox("Select Lot Column (Main)", df.columns)

#     if not carat_col:
#         carat_col = st.selectbox("Select Carat Column", df.columns)

#     if not lot_st:
#         lot_st = st.selectbox("Select Lot Column (Status File)", st_df.columns)

#     if not status_col:
#         status_col = st.selectbox("Select Status Column", st_df.columns)

#     if not lot_days:
#         lot_days = st.selectbox("Select Lot Column (Days File)", days_df.columns)

#     if not days_col:
#         days_col = st.selectbox("Select Days Column", days_df.columns)

#     # 🚀 GENERATE BUTTON
#     if st.button("🚀 Generate Final Excel"):

#         with st.spinner("Processing data... ⏳"):
#             try:

#                 # CLEAN CARAT
#                 df[carat_col] = (
#                     df[carat_col].astype(str)
#                     .str.replace("ct", "", regex=False)
#                     .str.replace(",", ".")
#                 )
#                 df[carat_col] = pd.to_numeric(df[carat_col], errors='coerce')

#                 # STATUS
#                 df["Status"] = df[lot_main].map(st_df.set_index(lot_st)[status_col])

#                 # DAYS
#                 df["No_of_Days"] = df[lot_main].map(days_df.set_index(lot_days)[days_col])

#                 # POSITION
#                 lot_index = df.columns.get_loc(lot_main)
#                 df.insert(lot_index + 1, "Status", df.pop("Status"))

#                 status_index = df.columns.get_loc("Status")
#                 df.insert(status_index + 1, "No_of_Days", df.pop("No_of_Days"))

#                 # SIZE GROUP
#                 size_series = df[carat_col].apply(assign_group)
#                 carat_index = df.columns.get_loc(carat_col)
#                 df.insert(carat_index + 1, "Size_Group", size_series)

#                 # SUCCESS
#                 st.success("✅ File Generated Successfully!")

#                 st.dataframe(df, use_container_width=True)

#                 # DOWNLOAD
#                 output_file = "final_output.xlsx"
#                 df.to_excel(output_file, index=False)

#                 with open(output_file, "rb") as f:
#                     st.download_button("📥 Download Excel", f, output_file)

#             except Exception as e:
#                 st.error("❌ Error while processing")
#                 st.write(e)



# import streamlit as st
# from openpyxl import load_workbook

# st.title("💎 Diamond Size Group Automation")

# file = st.file_uploader("Upload Excel", type=["xlsx"])

# if file:
#     wb = load_workbook(file)
#     ws = wb.active

#     # 🔍 Step 1: Find Carat column automatically
#     carat_col = None
#     for col in ws.iter_cols(1, ws.max_column):
#         if col[0].value is not None:
#             header = str(col[0].value).lower()
#             if "carat" in header or "ct" in header or "weight" in header:
#                 carat_col = col[0].column_letter
#                 break

#     if not carat_col:
#         st.error("❌ Carat column not found!")
#     else:
#         st.success(f"✅ Carat column found: {carat_col}")

#         # 📌 Step 2: Create Size Group column
#         new_col = ws.max_column + 1
#         size_col_letter = ws.cell(row=1, column=new_col).column_letter
#         ws.cell(row=1, column=new_col).value = "Size_Group"

#         # ⚙️ Step 3: Apply formula dynamically
#         for i in range(2, ws.max_row + 1):
#             ws[f"{size_col_letter}{i}"] = (
#                     f'=IF({carat_col}{i}<0.90,"Below 0.90",'
#                     f'IF({carat_col}{i}<=0.99,"0.90-0.99",'
#                     f'IF({carat_col}{i}<=1.10,"1.00-1.10",'
#                     f'IF({carat_col}{i}<=1.49,"1.11-1.49",'
#                     f'IF({carat_col}{i}<=1.99,"1.50-1.99",'
#                     f'IF({carat_col}{i}<=2.99,"2.00-2.99","3.00+"))))))'
#                 )

#         # 💾 Save file
#         output_file = "output.xlsx"
#         wb.save(output_file)

#         # 📥 Download button
#         with open(output_file, "rb") as f:
#             st.download_button("📥 Download File", f, file_name="output.xlsx")



# import streamlit as st
# from openpyxl import load_workbook

# st.title("💎 Diamond Pricing Automation")

# file = st.file_uploader("Upload Excel", type=["xlsx"])

# if file:
#     wb = load_workbook(file)
#     ws = wb.active

#     # 🔍 Detect columns
#     carat_col = None
#     cost_col = None

#     for col in ws.iter_cols(1, ws.max_column):
#         header = str(col[0].value).lower() if col[0].value else ""

#         if "carat" in header or "ct" in header:
#             carat_col = col[0].column_letter

#         if "old_price" in header or "price" in header:
#             cost_col = col[0].column_letter

#     if not carat_col:
#         st.error("❌ Carat column not found")
#     elif not cost_col:
#         st.error("❌ Old Price column not found")
#     else:
#         st.success(f"✅ Carat: {carat_col}, Cost: {cost_col}")

#         # 📌 Add new columns
#         size_col = ws.max_column + 1
#         updated_price_col = size_col + 1
#         diff_col = size_col + 2

#         size_letter = ws.cell(row=1, column=size_col).column_letter
#         updated_letter = ws.cell(row=1, column=updated_price_col).column_letter
#         diff_letter = ws.cell(row=1, column=diff_col).column_letter

#         ws[f"{size_letter}1"] = "Size_Group"
#         ws[f"{updated_letter}1"] = "Updated_Price"
#         ws[f"{diff_letter}1"] = "Difference_%"

#         # ⚙️ Apply formulas
#         for i in range(2, ws.max_row + 1):

#             # ✅ Size Group
#             ws[f"{size_letter}{i}"] = (
#                 f'=IF({carat_col}{i}<0.90,"Below 0.90",'
#                 f'IF({carat_col}{i}<=0.99,"0.90-0.99",'
#                 f'IF({carat_col}{i}<=1.10,"1.00-1.10",'
#                 f'IF({carat_col}{i}<=1.49,"1.11-1.49",'
#                 f'IF({carat_col}{i}<=1.99,"1.50-1.99",'
#                 f'IF({carat_col}{i}<=2.99,"2.00-2.99","3.00+"))))))'
#             )

#             # ✅ Updated Price → KEEP BLANK
#             ws[f"{updated_letter}{i}"] = ""

#             # ✅ Difference % formula
#             ws[f"{diff_letter}{i}"] = (
#                 f'=IF({updated_letter}{i}="","",ROUND(({cost_col}{i}-{updated_letter}{i})/{cost_col}{i}*100,2))'
#             )

#     # 💾 Save
#     output_file = "final_output.xlsx"
#     wb.save(output_file)

#     with open(output_file, "rb") as f:
#         st.download_button("📥 Download File", f, file_name="final_output.xlsx")
# import streamlit as st
# import pandas as pd
# from openpyxl import Workbook
# from openpyxl.utils import get_column_letter

# st.set_page_config(page_title="Diamond Tool", layout="wide")

# # ---------------- LOGIN ----------------
# def login_page():
#     st.title("Login Page")
#     username = st.text_input("Username")
#     password = st.text_input("Password", type="password")
#     if st.button("Login"):
#         if username == "admin" and password == "admin":
#             st.session_state.logged_in = True
#         else:
#             st.error("Invalid credentials")

# # ---------------- HOME ----------------
# def home_page():
#     st.title("Home")
#     option = st.radio("Select Option", ["Merge Files", "Repricing"])

#     if option == "Repricing":
#         repricing_page()

# # ---------------- HELPER ----------------
# def find_column(columns, keywords):
#     for col in columns:
#         for key in keywords:
#             if key in col.lower():
#                 return col
#     return None

# # ---------------- REPRICING ----------------
# def repricing_page():
#     st.header("Repricing Tool (FINAL PRO VERSION)")

#     file = st.file_uploader("Upload Excel", type=["xlsx"])

#     if file:
#         df = pd.read_excel(file)
#         cols = df.columns

#         carat_col = find_column(cols, ["crt", "carat", "cts", "weight"])
#         cost_col = find_column(cols, ["cost"])

#         if not carat_col:
#             carat_col = st.selectbox("Select Carat Column", cols)
#         if not cost_col:
#             cost_col = st.selectbox("Select Cost Column", cols)

#         # CLEAN $
#         df[cost_col] = df[cost_col].replace('[\$,]', '', regex=True).astype(float)

#         wb = Workbook()
#         ws = wb.active

#         headers = list(df.columns) + ["Size Group", "Updated Price", "Difference", "Total Cost", "Total Updated Value", "Profit/Loss"]
#         ws.append(headers)

#         for _, row in df.iterrows():
#             ws.append(list(row.values))

#         col_index = {col: i+1 for i, col in enumerate(df.columns)}

#         for r in range(2, len(df)+2):

#             carat_cell = f"{get_column_letter(col_index[carat_col])}{r}"
#             cost_cell = f"{get_column_letter(col_index[cost_col])}{r}"

#             size_col = len(df.columns) + 1
#             upd_col = len(df.columns) + 2
#             diff_col = len(df.columns) + 3
#             tc_col = len(df.columns) + 4
#             tu_col = len(df.columns) + 5
#             pl_col = len(df.columns) + 6

#             size_cell = f"{get_column_letter(size_col)}{r}"
#             upd_cell = f"{get_column_letter(upd_col)}{r}"
#             diff_cell = f"{get_column_letter(diff_col)}{r}"
#             tc_cell = f"{get_column_letter(tc_col)}{r}"
#             tu_cell = f"{get_column_letter(tu_col)}{r}"
#             pl_cell = f"{get_column_letter(pl_col)}{r}"

#             # ✅ FIXED SIZE GROUP (CLEAN & WORKING)
#             ws[size_cell] = f"=IF({carat_cell}<0.3,\"Other\",IF({carat_cell}<=0.49,\"0.30-0.49\",IF({carat_cell}<=0.59,\"0.50-0.59\",IF({carat_cell}<=0.69,\"0.60-0.69\",IF({carat_cell}<=0.79,\"0.70-0.79\",IF({carat_cell}<=0.89,\"0.80-0.89\",IF({carat_cell}<=0.99,\"0.90-0.99\",IF({carat_cell}<=1.10,\"1.00-1.10\",IF({carat_cell}<=1.49,\"1.11-1.49\",\"Other\"))))))))"

#             # ✅ 2 DECIMAL FIX USING ROUND
#             ws[diff_cell] = f"=ROUND(-(({cost_cell}-{upd_cell})/{cost_cell})*100,2)"
#             ws[tc_cell] = f"=ROUND({cost_cell}*{carat_cell},2)"
#             ws[tu_cell] = f"=ROUND({upd_cell}*{carat_cell},2)"
#             ws[pl_cell] = f"=ROUND({tu_cell}-{tc_cell},2)"

#         file_name = "FINAL_PRO_OUTPUT.xlsx"
#         wb.save(file_name)

#         with open(file_name, "rb") as f:
#             st.download_button("Download Final Excel", f, file_name=file_name)

# # ---------------- MAIN ----------------
# if "logged_in" not in st.session_state:
#     st.session_state.logged_in = False

# if not st.session_state.logged_in:
#     login_page()
# else:
#     home_page()
# import streamlit as st
# import pandas as pd
# from openpyxl import load_workbook

# st.title("💎 Diamond Size Group Automation")

# uploaded_file = st.file_uploader("Upload Excel File", type=["xlsx"])

# if uploaded_file:

#     df = pd.read_excel(uploaded_file)
#     st.write(df.head())

#     cts_column = st.selectbox("Select CTS Column", df.columns)

#     if st.button("Generate Size Group"):

#         # Convert CTS to numeric
#         df[cts_column] = pd.to_numeric(df[cts_column], errors='coerce')

#         temp_file = "temp.xlsx"
#         df.to_excel(temp_file, index=False)

#         wb = load_workbook(temp_file)

#         ws1 = wb.active
#         ws1.title = "Data"

#         # 🔥 Create lookup sheet
#         ws2 = wb.create_sheet(title="SizeMap")

#         size_data = [
#             (0.30, "0.30-0.49"),
#             (0.50, "0.50-0.59"),
#             (0.60, "0.60-0.69"),
#             (0.70, "0.70-0.79"),
#             (0.80, "0.80-0.89"),
#             (0.90, "0.90-0.99"),
#             (1.00, "1.00-1.10"),
#             (1.11, "1.11-1.49"),
#             (1.50, "1.50-1.59"),
#             (1.60, "1.60-1.99"),
#             (2.00, "2.00-2.10"),
#             (2.11, "2.11-2.49"),
#             (2.50, "2.50-2.59"),
#             (2.60, "2.60-2.99"),
#             (3.00, "3.00-3.10"),
#             (3.11, "3.11-3.49"),
#             (3.50, "3.50-3.59"),
#             (3.60, "3.60-3.99"),
#             (4.00, "4.00-4.10"),
#             (4.11, "4.11-4.49"),
#         ]

#         ws2.append(["Min CTS", "Size Group"])
#         for row in size_data:
#             ws2.append(row)

#         # Find CTS column index
#         col_index = list(df.columns).index(cts_column) + 1
#         new_col_index = col_index + 1

#         ws1.cell(row=1, column=new_col_index).value = "Size Group"

#         # Apply VLOOKUP formula
#         for row in range(2, ws1.max_row + 1):
#             formula = f'=VLOOKUP({chr(64+col_index)}{row},SizeMap!A:B,2,TRUE)'
#             ws1.cell(row=row, column=new_col_index).value = formula

#         wb.save("final_output.xlsx")

#         with open("final_output.xlsx", "rb") as f:
#             st.download_button("Download File", f, file_name="diamond_output.xlsx")
#


## Error


# Working






# import streamlit as st
# import pandas as pd
# from openpyxl import load_workbook
# from openpyxl.utils import get_column_letter

# st.title("💎 Diamond Size Group Automation")

# uploaded_file = st.file_uploader("Upload Excel File", type=["xlsx"])

# if uploaded_file:

#     df = pd.read_excel(uploaded_file)
#     st.write(df.head())

#     cts_column = st.selectbox("Select CTS Column", df.columns)

#     if st.button("Generate Full Automation"):

#         # Convert CTS to numeric
#         df[cts_column] = pd.to_numeric(df[cts_column], errors='coerce')

#         temp_file = "temp.xlsx"
#         df.to_excel(temp_file, index=False)

#         wb = load_workbook(temp_file)
#         ws1 = wb.active
#         ws1.title = "Data"

#         # 🔥 Create SizeMap
#         ws2 = wb.create_sheet(title="SizeMap")

#         size_data = [
#             (0.30, "0.30-0.49"), (0.50, "0.50-0.59"), (0.60, "0.60-0.69"),
#             (0.70, "0.70-0.79"), (0.80, "0.80-0.89"), (0.90, "0.90-0.99"),
#             (1.00, "1.00-1.10"), (1.11, "1.11-1.49"), (1.50, "1.50-1.59"),
#             (1.60, "1.60-1.99"), (2.00, "2.00-2.10"), (2.11, "2.11-2.49"),
#             (2.50, "2.50-2.59"), (2.60, "2.60-2.99"), (3.00, "3.00-3.10"),
#             (3.11, "3.11-3.49"), (3.50, "3.50-3.59"), (3.60, "3.60-3.99"),
#             (4.00, "4.00-4.10"), (4.11, "4.11-4.49"),
#         ]

#         ws2.append(["Min CTS", "Size Group"])
#         for row in size_data:
#             ws2.append(row)

#         # Column indexes
#         col_index = list(df.columns).index(cts_column) + 1

#         # 🔍 Find Cost column
#         cost_col_index = None
#         for i, col in enumerate(df.columns):
#             if "cost" in col.lower():
#                 cost_col_index = i + 1
#                 break

#         if cost_col_index is None:
#             st.error("Cost/CTS column not found!")
#             st.stop()

#         # ➕ Add Size Group
#         size_col = col_index + 1
#         ws1.cell(row=1, column=size_col).value = "Size Group"

#         for row in range(2, ws1.max_row + 1):
#             formula = f'=VLOOKUP({get_column_letter(col_index)}{row},SizeMap!A:B,2,TRUE)'
#             ws1.cell(row=row, column=size_col).value = formula

#         # ➕ Updated Price
#         updated_price_col = ws1.max_column + 1
#         ws1.cell(row=1, column=updated_price_col).value = "Updated Price"

#         # ➕ Difference %
#         diff_col = ws1.max_column + 1
#         ws1.cell(row=1, column=diff_col).value = "Difference (%)"

#         for row in range(2, ws1.max_row + 1):
#             cost_cell = f"{get_column_letter(cost_col_index)}{row}"
#             updated_cell = f"{get_column_letter(updated_price_col)}{row}"

#             formula = f"=-ROUND(({cost_cell}-{updated_cell})/{cost_cell}*100,2)"
#             ws1.cell(row=row, column=diff_col).value = formula

#         # 🔥 NEW 3 COLUMNS (ONLY CTS < 1)

#         cost_value_col = ws1.max_column + 1
#         updated_value_col = ws1.max_column + 2
#         profit_col = ws1.max_column + 3

#         ws1.cell(row=1, column=cost_value_col).value = "Cost Value"
#         ws1.cell(row=1, column=updated_value_col).value = "Updated Value"
#         ws1.cell(row=1, column=profit_col).value = "Profit/Loss"

#         for row in range(2, ws1.max_row + 1):

#             cts_cell = f"{get_column_letter(col_index)}{row}"
#             cost_cell = f"{get_column_letter(cost_col_index)}{row}"
#             updated_cell = f"{get_column_letter(updated_price_col)}{row}"

#             # Only apply if CTS < 1
#             cost_formula = f'=IF({cts_cell}<1,{cost_cell}*{cts_cell},"")'
#             updated_formula = f'=IF({cts_cell}<1,{updated_cell}*{cts_cell},"")'
#             profit_formula = f'=IF({cts_cell}<1,({updated_cell}*{cts_cell})-({cost_cell}*{cts_cell}),"")'

#             ws1.cell(row=row, column=cost_value_col).value = cost_formula
#             ws1.cell(row=row, column=updated_value_col).value = updated_formula
#             ws1.cell(row=row, column=profit_col).value = profit_formula

#         wb.save("final_output.xlsx")

#         with open("final_output.xlsx", "rb") as f:
#             st.download_button("Download File", f, file_name="diamond_output.xlsx")

# st.subheader("📊 Diamond Insights Dashboard")

# # 🔢 Total stones
# total_stones = len(df)

# # 🔹 Quality counts
# cvd_count = df[df['Quality'] == 'CVD'].shape[0] if 'Quality' in df.columns else 0
# hpht_count = df[df['Quality'] == 'HPHT'].shape[0] if 'Quality' in df.columns else 0

# # 📊 Metrics UI
# col1, col2, col3 = st.columns(3)

# col1.metric("Total Stones", total_stones)
# col2.metric("CVD Stones", cvd_count)
# col3.metric("HPHT Stones", hpht_count)

# # 🔷 Shape distribution
# if 'Shape' in df.columns:
#     st.subheader("💠 Shape Distribution")
#     shape_counts = df['Shape'].value_counts()
#     st.bar_chart(shape_counts)

# # 💰 Price insights
# if 'Price/CTS ($)' in df.columns and 'Cost/CTS ($)' in df.columns:
#     st.subheader("💰 Price Analysis")

#     avg_price = df['Price/CTS ($)'].mean()
#     avg_cost = df['Cost/CTS ($)'].mean()

#     col4, col5 = st.columns(2)
#     col4.metric("Avg Price/CTS", round(avg_price, 2))
#     col5.metric("Avg Cost/CTS", round(avg_cost, 2))

# # 🚨 Profit / Loss count (based on Difference % if exists)
# if 'Difference (%)' in df.columns:
#     profit_count = (df['Difference (%)'] > 0).sum()
#     loss_count = (df['Difference (%)'] < 0).sum()

#     st.subheader("📈 Profit vs Loss")
#     col6, col7 = st.columns(2)
#     col6.metric("Profit Stones", profit_count)
#     col7.metric("Loss Stones", loss_count)



# Final

# import streamlit as st
# import pandas as pd
# from openpyxl import load_workbook
# from openpyxl.utils import get_column_letter
# import matplotlib.pyplot as plt

# st.set_page_config(layout="wide")
# st.title("💎 Diamond Automation + Dashboard")

# uploaded_file = st.file_uploader("Upload Excel File", type=["xlsx"])

# if uploaded_file:

#     df = pd.read_excel(uploaded_file)

#     # ---------------- FILTERS ----------------
#     st.subheader("🔍 Filter Data")

#     col1, col2, col3 = st.columns(3)

#     if 'Shape' in df.columns:
#         shapes = col1.multiselect("Shape", df['Shape'].unique(), default=df['Shape'].unique())
#         df = df[df['Shape'].isin(shapes)]

#     if 'Color' in df.columns:
#         colors = col2.multiselect("Color", df['Color'].unique(), default=df['Color'].unique())
#         df = df[df['Color'].isin(colors)]

#     if 'Clarity' in df.columns:
#         clarity = col3.multiselect("Clarity", df['Clarity'].unique(), default=df['Clarity'].unique())
#         df = df[df['Clarity'].isin(clarity)]

#     st.subheader("📊 Data Preview")
#     st.dataframe(df.head())

#     # ---------------- DASHBOARD ----------------
#     st.subheader("📊 Dashboard")

#     total_stones = len(df)
#     cvd = df[df['Quality'] == 'CVD'].shape[0] if 'Quality' in df.columns else 0
#     hpht = df[df['Quality'] == 'HPHT'].shape[0] if 'Quality' in df.columns else 0

#     m1, m2, m3 = st.columns(3)
#     m1.metric("Total Stones", total_stones)
#     m2.metric("CVD", cvd)
#     m3.metric("HPHT", hpht)

#     # Pie Chart
#     if 'Quality' in df.columns:
#         st.subheader("🥧 CVD vs HPHT")
#         fig, ax = plt.subplots()
#         df['Quality'].value_counts().plot.pie(autopct='%1.1f%%', ax=ax)
#         st.pyplot(fig)

#     # Shape Chart
#     if 'Shape' in df.columns:
#         st.subheader("💠 Shape Distribution")
#         st.bar_chart(df['Shape'].value_counts())

#     # Profit Table
#     if 'Price/CTS ($)' in df.columns and 'Cost/CTS ($)' in df.columns:
#         df['Profit'] = df['Price/CTS ($)'] - df['Cost/CTS ($)']
#         st.subheader("🏆 Top Profit Stones")
#         st.dataframe(df.sort_values(by='Profit', ascending=False).head(10))

#     # Aging
#     if 'No of Days' in df.columns:
#         st.subheader("⏳ Aging > 100 Days")
#         old = df[df['No of Days'] > 100]
#         st.metric("Old Stones", len(old))
#         st.dataframe(old.head(10))

#     # ---------------- AUTOMATION ----------------
#     st.subheader("⚙️ Automation")

#     cts_column = st.selectbox("Select CTS Column", df.columns)

#     if st.button("🚀 Run Full Automation"):

#         df[cts_column] = pd.to_numeric(df[cts_column], errors='coerce')

#         temp_file = "temp.xlsx"
#         df.to_excel(temp_file, index=False)

#         wb = load_workbook(temp_file)
#         ws = wb.active

#         # SizeMap
#         ws2 = wb.create_sheet("SizeMap")

#         size_data = [
#             (0.30,"0.30-0.49"),(0.50,"0.50-0.59"),(0.60,"0.60-0.69"),
#             (0.70,"0.70-0.79"),(0.80,"0.80-0.89"),(0.90,"0.90-0.99"),
#             (1.00,"1.00-1.10"),(1.11,"1.11-1.49"),(1.50,"1.50-1.59"),
#             (1.60,"1.60-1.99"),(2.00,"2.00-2.10"),(2.11,"2.11-2.49"),
#             (2.50,"2.50-2.59"),(2.60,"2.60-2.99"),(3.00,"3.00-3.10"),
#             (3.11,"3.11-3.49"),(3.50,"3.50-3.59"),(3.60,"3.60-3.99"),
#             (4.00,"4.00-4.10"),(4.11,"4.11-4.49"),
#             (4.50,"4.50-4.59"),(4.60,"4.60-4.99"),
#             (5.00,"5.00-5.49"),(5.50,"5.50-5.99"),
#             (6.00,"6.00-6.99"),(7.00,"7.00-7.99"),
#             (8.00,"8.00-8.99"),(9.00,"9.00-9.99"),
#             (10.00,"10.00-10.99"),(11.00,"11.00-11.99"),
#             (12.00,"12.00-12.99"),(13.00,"13.00-13.99"),
#             (14.00,"14.00-14.99"),(15.00,"15.00-15.99"),
#             (16.00,"16.00-16.99"),(17.00,"17.00-17.99"),
#             (18.00,"18.00-18.99"),(19.00,"19.00-19.99"),
#             (20.00,"20.00-20.99"),(21.00,"21.00-21.99"),
#             (22.00,"22.00-22.99"),(23.00,"23.00-23.99"),
#             (24.00,"24.00-24.99"),(25.00,"25.00-25.99"),
#             (25.01,"High Carat Stone")   # 🔥 IMPORTANT
#         ]

#         ws2.append(["Min CTS","Size Group"])
#         for r in size_data:
#             ws2.append(r)

#         col_index = list(df.columns).index(cts_column) + 1

#         # Find Cost column
#         cost_col = None
#         for i, col in enumerate(df.columns):
#             if "cost" in col.lower():
#                 cost_col = i + 1
#                 break

#         if cost_col is None:
#             st.error("Cost column not found")
#             st.stop()

#         # Size Group
#         size_col = col_index + 1
#         ws.cell(row=1, column=size_col).value = "Size Group"

#         for r in range(2, ws.max_row+1):
#             ws.cell(row=r, column=size_col).value = f'=VLOOKUP({get_column_letter(col_index)}{r},SizeMap!A:B,2,TRUE)'

#         # Updated Price
#         up_col = ws.max_column + 1
#         ws.cell(row=1, column=up_col).value = "Updated Price"

#         # Difference
#         diff_col = ws.max_column + 1
#         ws.cell(row=1, column=diff_col).value = "Difference (%)"

#         for r in range(2, ws.max_row+1):
#             c = f"{get_column_letter(cost_col)}{r}"
#             u = f"{get_column_letter(up_col)}{r}"
#             ws.cell(row=r, column=diff_col).value = f"=-ROUND(({c}-{u})/{c}*100,2)"

#         # Pointer logic columns
#         cv = ws.max_column + 1
#         uv = ws.max_column + 2
#         pf = ws.max_column + 3

#         ws.cell(row=1, column=cv).value = "Cost Value"
#         ws.cell(row=1, column=uv).value = "Updated Value"
#         ws.cell(row=1, column=pf).value = "Profit/Loss"

#         for r in range(2, ws.max_row+1):
#             cts = f"{get_column_letter(col_index)}{r}"
#             cost = f"{get_column_letter(cost_col)}{r}"
#             upd = f"{get_column_letter(up_col)}{r}"

#             # Get CTS value directly
#             cts_value = ws.cell(row=r, column=col_index).value

#             if cts_value is not None and float(cts_value) < 1:

#                 cts = f"{get_column_letter(col_index)}{r}"
#                 cost = f"{get_column_letter(cost_col)}{r}"
#                 upd = f"{get_column_letter(up_col)}{r}"

#                 ws.cell(row=r, column=cv).value = f'={cost}*{cts}'
#                 ws.cell(row=r, column=uv).value = f'={upd}*{cts}'
#                 cost_value_cell = f"{get_column_letter(cv)}{r}"
#                 updated_value_cell = f"{get_column_letter(uv)}{r}"

#                 ws.cell(row=r, column=pf).value = f'={updated_value_cell}-{cost_value_cell}'

#             else:
#                 # Leave blank (NO formula at all)
#                 ws.cell(row=r, column=cv).value = ""
#                 ws.cell(row=r, column=uv).value = ""
#                 ws.cell(row=r, column=pf).value = ""

#         wb.save("diamond_output.xlsx")

#         with open("diamond_output.xlsx","rb") as f:
#             st.download_button("📥 Download Processed File", f, file_name="diamond_output.xlsx")
# import streamlit as st
# import pandas as pd
# from openpyxl import load_workbook
# from openpyxl.utils import get_column_letter
# import io

# st.set_page_config(page_title="💎 Diamond System", layout="wide")

# # ---------------- STYLE ----------------
# st.markdown("""
#     <style>
#     body {background-color: #0E1117; color: white;}
#     .title {
#         font-size: 35px;
#         font-weight: bold;
#         color: #00FFD1;
#         text-align: center;
#     }
#     .card {
#         padding: 20px;
#         border-radius: 10px;
#         background-color: #1c1f26;
#         margin: 10px 0;
#     }
#     </style>
# """, unsafe_allow_html=True)

# # ---------------- SESSION ----------------
# if "login" not in st.session_state:
#     st.session_state.login = False

# if "page" not in st.session_state:
#     st.session_state.page = "login"

# # ---------------- LOGIN PAGE ----------------
# if not st.session_state.login:

#     st.markdown('<p class="title">💎 Diamond System Login</p>', unsafe_allow_html=True)

#     col1, col2, col3 = st.columns([1,2,1])

#     with col2:
#         st.markdown('<div class="card">', unsafe_allow_html=True)

#         username = st.text_input("Username")
#         password = st.text_input("Password", type="password")

#         if st.button("Login"):
#             if username == "admin" and password == "1234":
#                 st.session_state.login = True
#                 st.session_state.page = "home"
#                 st.rerun()
#             else:
#                 st.error("Invalid Credentials ❌")

#         st.markdown('</div>', unsafe_allow_html=True)

# # ---------------- HOME PAGE ----------------
# elif st.session_state.page == "home":

#     st.markdown('<p class="title">🏠 Home Dashboard</p>', unsafe_allow_html=True)

#     col1, col2 = st.columns(2)

#     with col1:
#         st.markdown('<div class="card">', unsafe_allow_html=True)
#         st.subheader("📂 Merge Files")
#         st.info("Coming Soon 🚀")
#         st.markdown('</div>', unsafe_allow_html=True)

#     with col2:
#         st.markdown('<div class="card">', unsafe_allow_html=True)
#         st.subheader("⚙️ Formula Automation")

#         if st.button("Open Automation"):
#             st.session_state.page = "automation"
#             st.rerun()

#         st.markdown('</div>', unsafe_allow_html=True)

#     if st.button("Logout"):
#         st.session_state.login = False
#         st.session_state.page = "login"
#         st.rerun()

# # ---------------- AUTOMATION PAGE ----------------
# elif st.session_state.page == "automation":

#     st.markdown('<p class="title">💎 Diamond Automation</p>', unsafe_allow_html=True)

#     uploaded_file = st.file_uploader("Upload Excel File", type=["xlsx"])

#     if uploaded_file:

#         df = pd.read_excel(uploaded_file)
#         st.dataframe(df.head())

#         cts_column = st.selectbox("Select CTS Column", df.columns)

#         if st.button("🚀 Run Automation"):

#             df[cts_column] = pd.to_numeric(df[cts_column], errors='coerce')

#             temp_file = "temp.xlsx"
#             df.to_excel(temp_file, index=False)

#             wb = load_workbook(temp_file)
#             ws = wb.active

#             # ---------------- SIZE MAP ----------------
#             ws2 = wb.create_sheet("SizeMap")

#             size_data = [
#                 (0.30,"0.30-0.49"),(0.50,"0.50-0.59"),(0.60,"0.60-0.69"),
#                 (0.70,"0.70-0.79"),(0.80,"0.80-0.89"),(0.90,"0.90-0.99"),
#                 (1.00,"1.00-1.10"),(1.11,"1.11-1.49"),(1.50,"1.50-1.59"),
#                 (1.60,"1.60-1.99"),(2.00,"2.00-2.10"),(2.11,"2.11-2.49"),
#                 (2.50,"2.50-2.59"),(2.60,"2.60-2.99"),(3.00,"3.00-3.10"),
#                 (3.11,"3.11-3.49"),(3.50,"3.50-3.59"),(3.60,"3.60-3.99"),
#                 (4.00,"4.00-4.10"),(4.11,"4.11-4.49"),
#                 (4.50,"4.50-4.59"),(4.60,"4.60-4.99"),
#                 (5.00,"5.00-5.49"),(5.50,"5.50-5.99"),
#                 (6.00,"6.00-6.99"),(7.00,"7.00-7.99"),
#                 (8.00,"8.00-8.99"),(9.00,"9.00-9.99"),
#                 (10.00,"10.00-10.99"),(11.00,"11.00-11.99"),
#                 (12.00,"12.00-12.99"),(13.00,"13.00-13.99"),
#                 (14.00,"14.00-14.99"),(15.00,"15.00-15.99"),
#                 (16.00,"16.00-16.99"),(17.00,"17.00-17.99"),
#                 (18.00,"18.00-18.99"),(19.00,"19.00-19.99"),
#                 (20.00,"20.00-20.99"),(21.00,"21.00-21.99"),
#                 (22.00,"22.00-22.99"),(23.00,"23.00-23.99"),
#                 (24.00,"24.00-24.99"),(25.00,"25.00-25.99"),
#                 (25.01,"High Carat Stone")
#             ]

#             ws2.append(["Min CTS","Size Group"])
#             for r in size_data:
#                 ws2.append(r)

#             col_index = list(df.columns).index(cts_column) + 1

#             # Cost column detect
#             cost_col = None
#             for i, col in enumerate(df.columns):
#                 if "cost" in col.lower():
#                     cost_col = i + 1
#                     break

#             if cost_col is None:
#                 st.error("Cost column not found")
#                 st.stop()

#             # Size Group
#             size_col = col_index + 1
#             ws.cell(row=1, column=size_col).value = "Size Group"

#             for r in range(2, ws.max_row+1):
#                 ws.cell(row=r, column=size_col).value = f'=VLOOKUP({get_column_letter(col_index)}{r},SizeMap!A:B,2,TRUE)'

#             # Updated Price
#             up_col = ws.max_column + 1
#             ws.cell(row=1, column=up_col).value = "Updated Price"

#             # Difference
#             diff_col = ws.max_column + 1
#             ws.cell(row=1, column=diff_col).value = "Difference (%)"

#             for r in range(2, ws.max_row+1):
#                 c = f"{get_column_letter(cost_col)}{r}"
#                 u = f"{get_column_letter(up_col)}{r}"
#                 ws.cell(row=r, column=diff_col).value = f"=-ROUND(({c}-{u})/{c}*100,2)"

#             # Pointer logic
#             cv = ws.max_column + 1
#             uv = ws.max_column + 2
#             pf = ws.max_column + 3

#             ws.cell(row=1, column=cv).value = "Cost Value"
#             ws.cell(row=1, column=uv).value = "Updated Value"
#             ws.cell(row=1, column=pf).value = "Profit/Loss"

#             for r in range(2, ws.max_row+1):

#                 cts_value = ws.cell(row=r, column=col_index).value

#                 if cts_value is not None and float(cts_value) < 1:

#                     cts = f"{get_column_letter(col_index)}{r}"
#                     cost = f"{get_column_letter(cost_col)}{r}"
#                     upd = f"{get_column_letter(up_col)}{r}"

#                     ws.cell(row=r, column=cv).value = f'={cost}*{cts}'
#                     ws.cell(row=r, column=uv).value = f'={upd}*{cts}'
#                     ws.cell(row=r, column=pf).value = f'={get_column_letter(uv)}{r}-{get_column_letter(cv)}{r}'

#                 else:
#                     ws.cell(row=r, column=cv).value = ""
#                     ws.cell(row=r, column=uv).value = ""
#                     ws.cell(row=r, column=pf).value = ""

#             # ✅ FIXED DOWNLOAD
#             output = io.BytesIO()
#             wb.save(output)
#             output.seek(0)

#             st.download_button(
#                 label="📥 Download Excel File",
#                 data=output,
#                 file_name="diamond_output.xlsx",
#                 mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
#             )

#     col1, col2 = st.columns(2)

#     with col1:
#         if st.button("⬅ Back"):
#             st.session_state.page = "home"
#             st.rerun()

#     with col2:
#         if st.button("Logout"):
#             st.session_state.login = False
#             st.session_state.page = "login"
#             st.rerun()  












# import streamlit as st
# import pandas as pd
# from openpyxl import load_workbook
# import io

# # ---------------- CONFIG ----------------
# st.set_page_config(
#     page_title="GEMAI",
#     page_icon="💎",
#     layout="wide"
# )

# # ---------------- REMOVE SIDEBAR ----------------
# st.markdown("""
# <style>
# [data-testid="stSidebar"] {display: none;}
# </style>
# """, unsafe_allow_html=True)

# # ---------------- GLOBAL STYLE ----------------
# st.markdown("""
# <style>

# /* 🔥 Background */
# .stApp {
#     background: linear-gradient(135deg, #0f2027, #000000);
#     color: white;
# }

# /* 💎 Logo */
# .logo {
#     font-size: 30px;
#     font-weight: bold;
#     color: #00FFD1;
# }

# /* 🔝 Header */
# .header {
#     display:flex;
#     justify-content: space-between;
#     align-items:center;
# }

# /* 📦 Cards */
# .card {
#     padding: 30px;
#     border-radius: 12px;
#     background: #111;
#     border: 2px solid #1f2a30;
#     transition: 0.3s;
# }

# /* ✨ Hover */
# .card:hover {
#     border: 2px solid #00FFD1;
#     box-shadow: 0px 0px 25px #00FFD1;
#     transform: translateY(-5px);
# }

# /* 🚀 Buttons */
# .stButton>button {
#     background: #00FFD1;
#     color: black;
#     border-radius: 8px;
#     font-weight: bold;
#     padding: 10px 16px;
# }

# .stButton>button:hover {
#     background: #00c9a7;
# }

# /* 🔐 Login Box */
# .login-box {
#     max-width: 400px;
#     margin: auto;
#     margin-top: 120px;
#     padding: 30px;
#     border-radius: 12px;
#     background: #111;
#     border: 1px solid #00FFD1;
#     text-align: center;
# }

# /* 📏 Fix spacing */
# .block-container {
#     padding-top: 2rem;
# }

# </style>
# """, unsafe_allow_html=True)

# # ---------------- SESSION ----------------
# if "login" not in st.session_state:
#     st.session_state.login = False

# if "page" not in st.session_state:
#     st.session_state.page = "login"

# # ---------------- LOGIN PAGE ----------------
# if not st.session_state.login:

#     col1, col2, col3 = st.columns([2,3,2])

#     with col2:
#         st.markdown('<div class="login-box">', unsafe_allow_html=True)

#         st.markdown('<div class="logo">💎 GEMAI</div>', unsafe_allow_html=True)
#         st.caption("Smart Diamond Intelligence Platform")

#         username = st.text_input("Username")
#         password = st.text_input("Password", type="password")

#         if st.button("Login", use_container_width=True):
#             if username == "admin" and password == "1234":
#                 st.session_state.login = True
#                 st.session_state.page = "home"
#                 st.rerun()
#             else:
#                 st.error("Invalid Credentials")

#         st.markdown('</div>', unsafe_allow_html=True)

# # ---------------- HOME PAGE ----------------
# elif st.session_state.page == "home":

#     # Header
#     col1, col2 = st.columns([6,1])

#     with col1:
#         st.markdown('<div class="logo">💎 GEMAI</div>', unsafe_allow_html=True)

#     with col2:
#         if st.button("Logout"):
#             st.session_state.login = False
#             st.session_state.page = "login"
#             st.rerun()

#     st.write("")
#     st.subheader("🚀 Dashboard")

#     col1, col2 = st.columns(2)

#     # Merge Tool
#     with col1:
#         st.markdown('<div class="card">', unsafe_allow_html=True)
#         st.markdown("### 📂 Merge Files")
#         st.write("Combine multiple Excel files into one clean dataset.")

#         if st.button("Open Merge Tool"):
#             st.info("Coming Soon")

#         st.markdown('</div>', unsafe_allow_html=True)

#     # Automation Tool
#     with col2:
#         st.markdown('<div class="card">', unsafe_allow_html=True)
#         st.markdown("### ⚙️ Automation Engine")
#         st.write("Run diamond pricing and calculation automation.")

#         if st.button("Open Automation"):
#             st.session_state.page = "automation"
#             st.rerun()

#         st.markdown('</div>', unsafe_allow_html=True)

# # ---------------- AUTOMATION PAGE ----------------
# elif st.session_state.page == "automation":

#     col1, col2, col3 = st.columns([5,1,1])

#     with col1:
#         st.markdown('<div class="logo">⚙️ GEMAI Automation</div>', unsafe_allow_html=True)

#     with col2:
#         if st.button("Home"):
#             st.session_state.page = "home"
#             st.rerun()

#     with col3:
#         if st.button("Logout"):
#             st.session_state.login = False
#             st.session_state.page = "login"
#             st.rerun()

#     st.write("")

#     uploaded_file = st.file_uploader("Upload Excel File", type=["xlsx"])

#     if uploaded_file:

#         df = pd.read_excel(uploaded_file)
#         st.dataframe(df.head())

#         cts_column = st.selectbox("Select CTS Column", df.columns)

#         if st.button("Run Automation"):

#             df[cts_column] = pd.to_numeric(df[cts_column], errors='coerce')

#             temp_file = "temp.xlsx"
#             df.to_excel(temp_file, index=False)

#             wb = load_workbook(temp_file)

#             output = io.BytesIO()
#             wb.save(output)
#             output.seek(0)

#             st.download_button(
#                 "Download Excel",
#                 data=output,
#                 file_name="gemai_output.xlsx"
#             )




# import streamlit as st
# import pandas as pd
# from openpyxl import load_workbook
# from openpyxl.utils import get_column_letter
# import io

# # ---------------- CONFIG ----------------
# st.set_page_config(page_title="GEMAI", page_icon="💎", layout="wide")

# # ---------------- REMOVE SIDEBAR ----------------
# st.markdown("""
# <style>
# [data-testid="stSidebar"] {display: none;}
# </style>
# """, unsafe_allow_html=True)

# # ---------------- UI STYLE ----------------
# st.markdown("""
# <style>

# /* Background */
# .stApp {
#     background: linear-gradient(135deg, #0f2027, #000000);
#     color: white;
# }

# /* Logo */
# .logo {
#     font-size: 30px;
#     font-weight: bold;
#     color: #00FFD1;
# }

# /* Cards */
# .card {
#     padding: 25px;
#     border-radius: 12px;
#     background: #111;
#     border: 2px solid #1f2a30;
#     transition: 0.3s;
# }

# .card:hover {
#     border: 2px solid #00FFD1;
#     box-shadow: 0px 0px 20px #00FFD1;
# }

# /* Buttons */
# .stButton>button {
#     background: #00FFD1;
#     color: black;
#     border-radius: 8px;
#     font-weight: bold;
# }

# /* Login */
# .login-box {
#     max-width: 400px;
#     margin: auto;
#     margin-top: 120px;
#     padding: 30px;
#     background: #111;
#     border-radius: 12px;
#     border: 1px solid #00FFD1;
# }

# </style>
# """, unsafe_allow_html=True)

# # ---------------- SESSION ----------------
# if "login" not in st.session_state:
#     st.session_state.login = False

# if "page" not in st.session_state:
#     st.session_state.page = "login"

# # ---------------- LOGIN ----------------
# if not st.session_state.login:

#     col1, col2, col3 = st.columns([2,3,2])

#     with col2:
#         st.markdown('<div class="login-box">', unsafe_allow_html=True)
#         st.markdown('<div class="logo">💎 GEMAI</div>', unsafe_allow_html=True)

#         username = st.text_input("Username")
#         password = st.text_input("Password", type="password")

#         if st.button("Login", use_container_width=True):
#             if username == "admin" and password == "1234":
#                 st.session_state.login = True
#                 st.session_state.page = "home"
#                 st.rerun()
#             else:
#                 st.error("Invalid Credentials")

#         st.markdown('</div>', unsafe_allow_html=True)

# # ---------------- HOME ----------------
# elif st.session_state.page == "home":

#     col1, col2 = st.columns([6,1])

#     with col1:
#         st.markdown('<div class="logo">💎 GEMAI</div>', unsafe_allow_html=True)

#     with col2:
#         if st.button("Logout"):
#             st.session_state.login = False
#             st.session_state.page = "login"
#             st.rerun()

#     st.write("")
#     st.subheader("Dashboard")

#     col1, col2 = st.columns(2)

#     with col1:
#         st.markdown('<div class="card">', unsafe_allow_html=True)
#         st.subheader("📂 Merge Files")
#         st.info("Coming Soon")
#         st.markdown('</div>', unsafe_allow_html=True)

#     with col2:
#         st.markdown('<div class="card">', unsafe_allow_html=True)
#         st.subheader("⚙️ Automation")

#         if st.button("Open Automation"):
#             st.session_state.page = "automation"
#             st.rerun()

#         st.markdown('</div>', unsafe_allow_html=True)

# # ---------------- AUTOMATION ----------------
# elif st.session_state.page == "automation":

#     st.markdown('<div class="logo">⚙️ Diamond Automation</div>', unsafe_allow_html=True)

#     uploaded_file = st.file_uploader("Upload Excel File", type=["xlsx"])

#     if uploaded_file:

#         df = pd.read_excel(uploaded_file)
#         st.dataframe(df.head())

#         cts_column = st.selectbox("Select CTS Column", df.columns)

#         if st.button("🚀 Run Automation"):

#             df[cts_column] = pd.to_numeric(df[cts_column], errors='coerce')

#             temp_file = "temp.xlsx"
#             df.to_excel(temp_file, index=False)

#             wb = load_workbook(temp_file)
#             ws = wb.active

#             # SIZE MAP
#             ws2 = wb.create_sheet("SizeMap")

#             size_data = [
#                 (0.30,"0.30-0.49"),(0.50,"0.50-0.59"),(0.60,"0.60-0.69"),
#                 (0.70,"0.70-0.79"),(0.80,"0.80-0.89"),(0.90,"0.90-0.99"),
#                 (1.00,"1.00-1.10"),(1.11,"1.11-1.49"),(1.50,"1.50-1.59"),
#                 (1.60,"1.60-1.99"),(2.00,"2.00-2.10"),(2.11,"2.11-2.49"),
#                 (2.50,"2.50-2.59"),(2.60,"2.60-2.99"),(3.00,"3.00-3.10"),
#                 (3.11,"3.11-3.49"),(3.50,"3.50-3.59"),(3.60,"3.60-3.99"),
#                 (4.00,"4.00-4.10"),(4.11,"4.11-4.49"),
#                 (4.50,"4.50-4.59"),(4.60,"4.60-4.99"),
#                 (5.00,"5.00-5.49"),(5.50,"5.50-5.99"),
#                 (6.00,"6.00-6.99"),(7.00,"7.00-7.99"),
#                 (8.00,"8.00-8.99"),(9.00,"9.00-9.99"),
#                 (10.00,"10.00-10.99"),(11.00,"11.00-11.99"),
#                 (12.00,"12.00-12.99"),(13.00,"13.00-13.99"),
#                 (14.00,"14.00-14.99"),(15.00,"15.00-15.99"),
#                 (16.00,"16.00-16.99"),(17.00,"17.00-17.99"),
#                 (18.00,"18.00-18.99"),(19.00,"19.00-19.99"),
#                 (20.00,"20.00-20.99"),(21.00,"21.00-21.99"),
#                 (22.00,"22.00-22.99"),(23.00,"23.00-23.99"),
#                 (24.00,"24.00-24.99"),(25.00,"25.00-25.99"),
#                 (25.01,"High Carat Stone")
#             ]

#             ws2.append(["Min CTS","Size Group"])
#             for r in size_data:
#                 ws2.append(r)

#             col_index = list(df.columns).index(cts_column) + 1

#             cost_col = None
#             for i, col in enumerate(df.columns):
#                 if "cost" in col.lower():
#                     cost_col = i + 1
#                     break

#             if cost_col is None:
#                 st.error("Cost column not found")
#                 st.stop()

#             size_col = col_index + 1
#             ws.cell(row=1, column=size_col).value = "Size Group"

#             for r in range(2, ws.max_row+1):
#                 ws.cell(row=r, column=size_col).value = f'=VLOOKUP({get_column_letter(col_index)}{r},SizeMap!A:B,2,TRUE)'

#             up_col = ws.max_column + 1
#             ws.cell(row=1, column=up_col).value = "Updated Price"

#             diff_col = ws.max_column + 1
#             ws.cell(row=1, column=diff_col).value = "Difference (%)"

#             for r in range(2, ws.max_row+1):
#                 c = f"{get_column_letter(cost_col)}{r}"
#                 u = f"{get_column_letter(up_col)}{r}"
#                 ws.cell(row=r, column=diff_col).value = f"=-ROUND(({c}-{u})/{c}*100,2)"

#             cv = ws.max_column + 1
#             uv = ws.max_column + 2
#             pf = ws.max_column + 3

#             ws.cell(row=1, column=cv).value = "Cost Value"
#             ws.cell(row=1, column=uv).value = "Updated Value"
#             ws.cell(row=1, column=pf).value = "Profit/Loss"

#             for r in range(2, ws.max_row+1):

#                 cts_value = ws.cell(row=r, column=col_index).value

#                 if cts_value is not None and float(cts_value) < 1:

#                     cts = f"{get_column_letter(col_index)}{r}"
#                     cost = f"{get_column_letter(cost_col)}{r}"
#                     upd = f"{get_column_letter(up_col)}{r}"

#                     ws.cell(row=r, column=cv).value = f'={cost}*{cts}'
#                     ws.cell(row=r, column=uv).value = f'={upd}*{cts}'
#                     ws.cell(row=r, column=pf).value = f'={get_column_letter(uv)}{r}-{get_column_letter(cv)}{r}'

#             output = io.BytesIO()
#             wb.save(output)
#             output.seek(0)

#             st.download_button(
#                 label="📥 Download Excel File",
#                 data=output,
#                 file_name="diamond_output.xlsx"
#             )

#     col1, col2 = st.columns(2)

#     with col1:
#         if st.button("⬅ Back"):
#             st.session_state.page = "home"
#             st.rerun()

#     with col2:
#         if st.button("Logout"):
#             st.session_state.login = False
#             st.session_state.page = "login"
#             st.rerun()





# Automation and merge

# import streamlit as st
# import pandas as pd
# from openpyxl import load_workbook
# from openpyxl.utils import get_column_letter
# import io

# # ---------------- CONFIG ----------------
# st.set_page_config(page_title="GEMAI", page_icon="💎", layout="wide")

# # ---------------- REMOVE SIDEBAR ----------------
# st.markdown("""
# <style>
# [data-testid="stSidebar"] {display: none;}
# </style>
# """, unsafe_allow_html=True)

# # ---------------- UI STYLE ----------------
# st.markdown("""
# <style>

# /* Background */
# .stApp {
#     background: linear-gradient(135deg, #0f2027, #000000);
#     color: white;
# }

# /* Logo */
# .logo {
#     font-size: 30px;
#     font-weight: bold;
#     color: #00FFD1;
# }

# /* Cards */
# .card {
#     padding: 25px;
#     border-radius: 12px;
#     background: #111;
#     border: 2px solid #1f2a30;
#     transition: 0.3s;
# }

# .card:hover {
#     border: 2px solid #00FFD1;
#     box-shadow: 0px 0px 20px #00FFD1;
# }

# /* Buttons */
# .stButton>button {
#     background: #00FFD1;
#     color: black;
#     border-radius: 8px;
#     font-weight: bold;
# }

# /* Login */
# .login-box {
#     max-width: 400px;
#     margin: auto;
#     margin-top: 120px;
#     padding: 30px;
#     background: #111;
#     border-radius: 12px;
#     border: 1px solid #00FFD1;
# }

# </style>
# """, unsafe_allow_html=True)

# # ---------------- SESSION ----------------
# if "login" not in st.session_state:
#     st.session_state.login = False

# if "page" not in st.session_state:
#     st.session_state.page = "login"

# # ---------------- LOGIN ----------------
# if not st.session_state.login:

#     col1, col2, col3 = st.columns([2,3,2])

#     with col2:
#         st.markdown('<div class="login-box">', unsafe_allow_html=True)
#         st.markdown('<div class="logo">💎 GEMAI</div>', unsafe_allow_html=True)

#         username = st.text_input("Username")
#         password = st.text_input("Password", type="password")

#         if st.button("Login", use_container_width=True):
#             if username == "admin" and password == "1234":
#                 st.session_state.login = True
#                 st.session_state.page = "home"
#                 st.rerun()
#             else:
#                 st.error("Invalid Credentials")

#         st.markdown('</div>', unsafe_allow_html=True)

# # ---------------- HOME ----------------
# elif st.session_state.page == "home":

#     col1, col2 = st.columns([6,1])

#     with col1:
#         st.markdown('<div class="logo">💎 GEMAI</div>', unsafe_allow_html=True)

#     with col2:
#         if st.button("Logout"):
#             st.session_state.login = False
#             st.session_state.page = "login"
#             st.rerun()

#     st.write("")
#     st.subheader("Dashboard")

#     col1, col2 = st.columns(2)

#     with col1:
#         st.markdown('<div class="card">', unsafe_allow_html=True)
#         st.subheader("📂 Merge Files")
#         st.info("Coming Soon")
#         st.markdown('</div>', unsafe_allow_html=True)

#     with col2:
#         st.markdown('<div class="card">', unsafe_allow_html=True)
#         st.subheader("⚙️ Automation")

#         if st.button("Open Automation"):
#             st.session_state.page = "automation"
#             st.rerun()

#         st.markdown('</div>', unsafe_allow_html=True)

# # ---------------- AUTOMATION ----------------
# elif st.session_state.page == "automation":

#     st.markdown('<div class="logo">⚙️ Diamond Automation</div>', unsafe_allow_html=True)

#     uploaded_file = st.file_uploader("Upload Excel File", type=["xlsx"])

#     if uploaded_file:

#         df = pd.read_excel(uploaded_file)

#         # ✅ NEW: Detect LOT column (instead of Stock ID)
#         lot_column = None
#         for col in df.columns:
#             if "lot" in col.lower():
#                 lot_column = col
#                 break

#         if lot_column is None:
#             st.error("Lot column not found")
#             st.stop()

#         st.success(f"Using '{lot_column}' as identifier")

#         st.dataframe(df.head())

#         cts_column = st.selectbox("Select CTS Column", df.columns)

#         if st.button("🚀 Run Automation"):

#             df[cts_column] = pd.to_numeric(df[cts_column], errors='coerce')

#             temp_file = "temp.xlsx"
#             df.to_excel(temp_file, index=False)

#             wb = load_workbook(temp_file)
#             ws = wb.active

#             # SIZE MAP
#             ws2 = wb.create_sheet("SizeMap")

#             size_data = [
#                 (0.30,"0.30-0.49"),(0.50,"0.50-0.59"),(0.60,"0.60-0.69"),
#                 (0.70,"0.70-0.79"),(0.80,"0.80-0.89"),(0.90,"0.90-0.99"),
#                 (1.00,"1.00-1.10"),(1.11,"1.11-1.49"),(1.50,"1.50-1.59"),
#                 (1.60,"1.60-1.99"),(2.00,"2.00-2.10"),(2.11,"2.11-2.49"),
#                 (2.50,"2.50-2.59"),(2.60,"2.60-2.99"),(3.00,"3.00-3.10"),
#                 (3.11,"3.11-3.49"),(3.50,"3.50-3.59"),(3.60,"3.60-3.99"),
#                 (4.00,"4.00-4.10"),(4.11,"4.11-4.49"),
#                 (4.50,"4.50-4.59"),(4.60,"4.60-4.99"),
#                 (5.00,"5.00-5.49"),(5.50,"5.50-5.99"),
#                 (6.00,"6.00-6.99"),(7.00,"7.00-7.99"),
#                 (8.00,"8.00-8.99"),(9.00,"9.00-9.99"),
#                 (10.00,"10.00-10.99"),(11.00,"11.00-11.99"),
#                 (12.00,"12.00-12.99"),(13.00,"13.00-13.99"),
#                 (14.00,"14.00-14.99"),(15.00,"15.00-15.99"),
#                 (16.00,"16.00-16.99"),(17.00,"17.00-17.99"),
#                 (18.00,"18.00-18.99"),(19.00,"19.00-19.99"),
#                 (20.00,"20.00-20.99"),(21.00,"21.00-21.99"),
#                 (22.00,"22.00-22.99"),(23.00,"23.00-23.99"),
#                 (24.00,"24.00-24.99"),(25.00,"25.00-25.99"),
#                 (25.01,"High Carat Stone")
#             ]

#             ws2.append(["Min CTS","Size Group"])
#             for r in size_data:
#                 ws2.append(r)

#             col_index = list(df.columns).index(cts_column) + 1

#             cost_col = None
#             for i, col in enumerate(df.columns):
#                 if "cost" in col.lower():
#                     cost_col = i + 1
#                     break

#             if cost_col is None:
#                 st.error("Cost column not found")
#                 st.stop()

#             size_col = col_index + 1
#             ws.cell(row=1, column=size_col).value = "Size Group"

#             for r in range(2, ws.max_row+1):
#                 ws.cell(row=r, column=size_col).value = f'=VLOOKUP({get_column_letter(col_index)}{r},SizeMap!A:B,2,TRUE)'

#             up_col = ws.max_column + 1
#             ws.cell(row=1, column=up_col).value = "Updated Price"

#             diff_col = ws.max_column + 1
#             ws.cell(row=1, column=diff_col).value = "Difference (%)"

#             for r in range(2, ws.max_row+1):
#                 c = f"{get_column_letter(cost_col)}{r}"
#                 u = f"{get_column_letter(up_col)}{r}"
#                 ws.cell(row=r, column=diff_col).value = f"=-ROUND(({c}-{u})/{c}%,2)"

#             cv = ws.max_column + 1
#             uv = ws.max_column + 2
#             pf = ws.max_column + 3

#             ws.cell(row=1, column=cv).value = "Cost Value"
#             ws.cell(row=1, column=uv).value = "Updated Value"
#             ws.cell(row=1, column=pf).value = "Profit/Loss"

#             for r in range(2, ws.max_row+1):

#                 cts_value = ws.cell(row=r, column=col_index).value

#                 if cts_value is not None and float(cts_value) < 1:

#                     cts = f"{get_column_letter(col_index)}{r}"
#                     cost = f"{get_column_letter(cost_col)}{r}"
#                     upd = f"{get_column_letter(up_col)}{r}"

#                     ws.cell(row=r, column=cv).value = f'={cost}*{cts}'
#                     ws.cell(row=r, column=uv).value = f'={upd}*{cts}'
#                     ws.cell(row=r, column=pf).value = f'={get_column_letter(uv)}{r}-{get_column_letter(cv)}{r}'

#             output = io.BytesIO()
#             wb.save(output)
#             output.seek(0)

#             st.download_button(
#                 label="📥 Download Excel File",
#                 data=output,
#                 file_name="diamond_output.xlsx"
#             )

#     col1, col2 = st.columns(2)

#     with col1:
#         if st.button("⬅ Back"):
#             st.session_state.page = "home"
#             st.rerun()

#     with col2:
#         if st.button("Logout"):
#             st.session_state.login = False
#             st.session_state.page = "login"
#             st.rerun()


# import streamlit as st
# import pandas as pd
# from openpyxl import load_workbook
# from openpyxl.utils import get_column_letter
# import io

# # ---------------- CONFIG ----------------
# st.set_page_config(page_title="GEMAI", page_icon="💎", layout="wide")

# # ---------------- REMOVE SIDEBAR ----------------
# st.markdown("""
# <style>
# [data-testid="stSidebar"] {display: none;}
# </style>
# """, unsafe_allow_html=True)

# # ---------------- UI STYLE ----------------
# st.markdown("""
# <style>

# /* Background */
# .stApp {
#     background: linear-gradient(135deg, #0f2027, #000000);
#     color: white;
# }

# /* Logo */
# .logo {
#     font-size: 30px;
#     font-weight: bold;
#     color: #00FFD1;
# }

# /* Cards */
# .card {
#     padding: 25px;
#     border-radius: 12px;
#     background: #111;
#     border: 2px solid #1f2a30;
#     transition: 0.3s;
# }

# .card:hover {
#     border: 2px solid #00FFD1;
#     box-shadow: 0px 0px 20px #00FFD1;
# }

# /* Buttons */
# .stButton>button {
#     background: #00FFD1;
#     color: black;
#     border-radius: 8px;
#     font-weight: bold;
# }

# /* Login */
# .login-box {
#     max-width: 400px;
#     margin: auto;
#     margin-top: 120px;
#     padding: 30px;
#     background: #111;
#     border-radius: 12px;
#     border: 1px solid #00FFD1;
# }

# </style>
# """, unsafe_allow_html=True)

# # ---------------- SESSION ----------------
# if "login" not in st.session_state:
#     st.session_state.login = False

# if "page" not in st.session_state:
#     st.session_state.page = "login"

# # ---------------- LOGIN ----------------
# if not st.session_state.login:

#     col1, col2, col3 = st.columns([2,3,2])

#     with col2:
#         st.markdown('<div class="login-box">', unsafe_allow_html=True)
#         st.markdown('<div class="logo">💎 GEMAI</div>', unsafe_allow_html=True)

#         username = st.text_input("Username")
#         password = st.text_input("Password", type="password")

#         if st.button("Login", use_container_width=True):
#             if username == "admin" and password == "1234":
#                 st.session_state.login = True
#                 st.session_state.page = "home"
#                 st.rerun()
#             else:
#                 st.error("Invalid Credentials")

#         st.markdown('</div>', unsafe_allow_html=True)

# # ---------------- HOME ----------------
# elif st.session_state.page == "home":

#     col1, col2 = st.columns([6,1])

#     with col1:
#         st.markdown('<div class="logo">💎 GEMAI</div>', unsafe_allow_html=True)

#     with col2:
#         if st.button("Logout"):
#             st.session_state.login = False
#             st.session_state.page = "login"
#             st.rerun()

#     st.write("")
#     st.subheader("Dashboard")

#     col1, col2 = st.columns(2)

#     with col1:
#         st.markdown('<div class="card">', unsafe_allow_html=True)
#         st.subheader("📂 Merge Files")

#         if st.button("Open Merge Files"):
#             st.session_state.page = "merge"
#             st.rerun()

#         st.markdown('</div>', unsafe_allow_html=True)

#     with col2:
#         st.markdown('<div class="card">', unsafe_allow_html=True)
#         st.subheader("⚙️ Automation")

#         if st.button("Open Automation"):
#             st.session_state.page = "automation"
#             st.rerun()

#         st.markdown('</div>', unsafe_allow_html=True)

# # ---------------- MERGE FILES ----------------
# elif st.session_state.page == "merge":

#     st.markdown('<div class="logo">📂 Merge Files</div>', unsafe_allow_html=True)

#     file_main = st.file_uploader("Upload Main File", type=["xlsx"])
#     file_status = st.file_uploader("Upload Status File", type=["xlsx"])
#     file_days = st.file_uploader("Upload Days File", type=["xlsx"])

#     if file_main and file_status and file_days:

#         df_main = pd.read_excel(file_main)
#         df_status = pd.read_excel(file_status)
#         df_days = pd.read_excel(file_days)

#         # detect LOT column
#         lot_col = None
#         for col in df_main.columns:
#             if "lot" in col.lower():
#                 lot_col = col
#                 break

#         if lot_col is None:
#             st.error("Lot column not found")
#             st.stop()

#         temp_file = "merge_temp.xlsx"

#         with pd.ExcelWriter(temp_file, engine="openpyxl") as writer:
#             df_main.to_excel(writer, sheet_name="Main", index=False)
#             df_status.to_excel(writer, sheet_name="Status", index=False)
#             df_days.to_excel(writer, sheet_name="Days", index=False)

#         wb = load_workbook(temp_file)
#         ws = wb["Main"]

#         lot_index = list(df_main.columns).index(lot_col) + 1
#         lot_letter = get_column_letter(lot_index)

#         status_col = ws.max_column + 1
#         ws.cell(row=1, column=status_col).value = "Status"

#         days_col = ws.max_column + 1
#         ws.cell(row=1, column=days_col).value = "No_of_Days"

#         for r in range(2, ws.max_row + 1):

#             ws.cell(row=r, column=status_col).value = \
#                 f'=IFERROR(VLOOKUP({lot_letter}{r},Status!A:B,2,FALSE),"Not Found")'

#             ws.cell(row=r, column=days_col).value = \
#                 f'=IFERROR(VLOOKUP({lot_letter}{r},Days!A:B,2,FALSE),"Not Found")'

#         output = io.BytesIO()
#         wb.save(output)
#         output.seek(0)

#         st.download_button(
#             label="📥 Download Merged File",
#             data=output,
#             file_name="diamond_merged.xlsx"
#         )

#     col1, col2 = st.columns(2)

#     with col1:
#         if st.button("⬅ Back"):
#             st.session_state.page = "home"
#             st.rerun()

#     with col2:
#         if st.button("Logout"):
#             st.session_state.login = False
#             st.session_state.page = "login"
#             st.rerun()

# # ---------------- AUTOMATION ----------------
# elif st.session_state.page == "automation":

#     st.markdown('<div class="logo">⚙️ Diamond Automation</div>', unsafe_allow_html=True)

#     uploaded_file = st.file_uploader("Upload Excel File", type=["xlsx"])

#     if uploaded_file:

#         df = pd.read_excel(uploaded_file)

#         # ✅ ORIGINAL LOGIC UNCHANGED
#         lot_column = None
#         for col in df.columns:
#             if "lot" in col.lower():
#                 lot_column = col
#                 break

#         if lot_column is None:
#             st.error("Lot column not found")
#             st.stop()

#         st.success(f"Using '{lot_column}' as identifier")

#         st.dataframe(df.head())

#         cts_column = st.selectbox("Select CTS Column", df.columns)

#         if st.button("🚀 Run Automation"):

#             df[cts_column] = pd.to_numeric(df[cts_column], errors='coerce')

#             temp_file = "temp.xlsx"
#             df.to_excel(temp_file, index=False)

#             wb = load_workbook(temp_file)
#             ws = wb.active

#             # SIZE MAP
#             ws2 = wb.create_sheet("SizeMap")

#             size_data = [
#                 (0.30,"0.30-0.49"),(0.50,"0.50-0.59"),(0.60,"0.60-0.69"),
#                 (0.70,"0.70-0.79"),(0.80,"0.80-0.89"),(0.90,"0.90-0.99"),
#                 (1.00,"1.00-1.10"),(1.11,"1.11-1.49"),(1.50,"1.50-1.59"),
#                 (1.60,"1.60-1.99"),(2.00,"2.00-2.10"),(2.11,"2.11-2.49"),
#                 (2.50,"2.50-2.59"),(2.60,"2.60-2.99"),(3.00,"3.00-3.10"),
#                 (3.11,"3.11-3.49"),(3.50,"3.50-3.59"),(3.60,"3.60-3.99"),
#                 (4.00,"4.00-4.10"),(4.11,"4.11-4.49"),
#                 (4.50,"4.50-4.59"),(4.60,"4.60-4.99"),
#                 (5.00,"5.00-5.49"),(5.50,"5.50-5.99"),
#                 (6.00,"6.00-6.99"),(7.00,"7.00-7.99"),
#                 (8.00,"8.00-8.99"),(9.00,"9.00-9.99"),
#                 (10.00,"10.00-10.99"),(11.00,"11.00-11.99"),
#                 (12.00,"12.00-12.99"),(13.00,"13.00-13.99"),
#                 (14.00,"14.00-14.99"),(15.00,"15.00-15.99"),
#                 (16.00,"16.00-16.99"),(17.00,"17.00-17.99"),
#                 (18.00,"18.00-18.99"),(19.00,"19.00-19.99"),
#                 (20.00,"20.00-20.99"),(21.00,"21.00-21.99"),
#                 (22.00,"22.00-22.99"),(23.00,"23.00-23.99"),
#                 (24.00,"24.00-24.99"),(25.00,"25.00-25.99"),
#                 (25.01,"High Carat Stone")
#             ]

#             ws2.append(["Min CTS","Size Group"])
#             for r in size_data:
#                 ws2.append(r)

#             col_index = list(df.columns).index(cts_column) + 1

#             cost_col = None
#             for i, col in enumerate(df.columns):
#                 if "cost" in col.lower():
#                     cost_col = i + 1
#                     break

#             if cost_col is None:
#                 st.error("Cost column not found")
#                 st.stop()

#             size_col = col_index + 1
#             ws.cell(row=1, column=size_col).value = "Size Group"

#             for r in range(2, ws.max_row+1):
#                 ws.cell(row=r, column=size_col).value = f'=VLOOKUP({get_column_letter(col_index)}{r},SizeMap!A:B,2,TRUE)'

#             up_col = ws.max_column + 1
#             ws.cell(row=1, column=up_col).value = "Updated Price"

#             diff_col = ws.max_column + 1
#             ws.cell(row=1, column=diff_col).value = "Difference (%)"

#             for r in range(2, ws.max_row+1):
#                 c = f"{get_column_letter(cost_col)}{r}"
#                 u = f"{get_column_letter(up_col)}{r}"
#                 ws.cell(row=r, column=diff_col).value = f"=-ROUND(({c}-{u})/{c}%,2)"

#             cv = ws.max_column + 1
#             uv = ws.max_column + 2
#             pf = ws.max_column + 3

#             ws.cell(row=1, column=cv).value = "Cost Value"
#             ws.cell(row=1, column=uv).value = "Updated Value"
#             ws.cell(row=1, column=pf).value = "Profit/Loss"

#             for r in range(2, ws.max_row+1):

#                 cts_value = ws.cell(row=r, column=col_index).value

#                 if cts_value is not None and float(cts_value) < 1:

#                     cts = f"{get_column_letter(col_index)}{r}"
#                     cost = f"{get_column_letter(cost_col)}{r}"
#                     upd = f"{get_column_letter(up_col)}{r}"

#                     ws.cell(row=r, column=cv).value = f'={cost}*{cts}'
#                     ws.cell(row=r, column=uv).value = f'={upd}*{cts}'
#                     ws.cell(row=r, column=pf).value = f'={get_column_letter(uv)}{r}-{get_column_letter(cv)}{r}'

#             output = io.BytesIO()
#             wb.save(output)
#             output.seek(0)

#             st.download_button(
#                 label="📥 Download Excel File",
#                 data=output,
#                 file_name="diamond_output.xlsx"
#             )

#     col1, col2 = st.columns(2)

#     with col1:
#         if st.button("⬅ Back"):
#             st.session_state.page = "home"
#             st.rerun()

#     with col2:
#         if st.button("Logout"):
#             st.session_state.login = False
#             st.session_state.page = "login"
#             st.rerun()







# Merging wroks corretly

# import streamlit as st
# import pandas as pd
# from openpyxl import load_workbook
# from openpyxl.utils import get_column_letter
# import io

# # ---------------- CONFIG ----------------
# st.set_page_config(page_title="GEMAI", page_icon="💎", layout="wide")

# # ---------------- REMOVE SIDEBAR ----------------
# st.markdown("""
# <style>
# [data-testid="stSidebar"] {display: none;}
# </style>
# """, unsafe_allow_html=True)

# # ---------------- UI STYLE ----------------
# st.markdown("""
# <style>
# .stApp {
#     background: linear-gradient(135deg, #0f2027, #000000);
#     color: white;
# }
# .logo {
#     font-size: 30px;
#     font-weight: bold;
#     color: #00FFD1;
# }
# .card {
#     padding: 25px;
#     border-radius: 12px;
#     background: #111;
#     border: 2px solid #1f2a30;
# }
# .stButton>button {
#     background: #00FFD1;
#     color: black;
#     border-radius: 8px;
#     font-weight: bold;
# }
# .login-box {
#     max-width: 400px;
#     margin: auto;
#     margin-top: 120px;
#     padding: 30px;
#     background: #111;
#     border-radius: 12px;
#     border: 1px solid #00FFD1;
# }
# </style>
# """, unsafe_allow_html=True)

# # ---------------- SESSION ----------------
# if "login" not in st.session_state:
#     st.session_state.login = False

# if "page" not in st.session_state:
#     st.session_state.page = "login"

# # ---------------- LOGIN ----------------
# if not st.session_state.login:

#     col1, col2, col3 = st.columns([2,3,2])

#     with col2:
#         st.markdown('<div class="login-box">', unsafe_allow_html=True)
#         st.markdown('<div class="logo">💎 GEMAI</div>', unsafe_allow_html=True)

#         username = st.text_input("Username")
#         password = st.text_input("Password", type="password")

#         if st.button("Login", use_container_width=True):
#             if username == "admin" and password == "1234":
#                 st.session_state.login = True
#                 st.session_state.page = "home"
#                 st.rerun()
#             else:
#                 st.error("Invalid Credentials")

#         st.markdown('</div>', unsafe_allow_html=True)

# # ---------------- HOME ----------------
# elif st.session_state.page == "home":

#     col1, col2 = st.columns([6,1])

#     with col1:
#         st.markdown('<div class="logo">💎 GEMAI</div>', unsafe_allow_html=True)

#     with col2:
#         if st.button("Logout"):
#             st.session_state.login = False
#             st.session_state.page = "login"
#             st.rerun()

#     st.subheader("Dashboard")

#     col1, col2, col3 = st.columns(3)

#     with col1:
#         st.markdown('<div class="card">', unsafe_allow_html=True)
#         st.subheader("📂 Merge Files")
#         if st.button("Open Merge Files"):
#             st.session_state.page = "merge"
#             st.rerun()
#         st.markdown('</div>', unsafe_allow_html=True)

#     with col2:
#         st.markdown('<div class="card">', unsafe_allow_html=True)
#         st.subheader("⚙️ Automation")
#         if st.button("Open Automation"):
#             st.session_state.page = "automation"
#             st.rerun()
#         st.markdown('</div>', unsafe_allow_html=True)

#     with col3:
#         st.markdown('<div class="card">', unsafe_allow_html=True)
#         st.subheader("⚡ Smart Lookup")
#         if st.button("Open Smart Lookup"):
#             st.session_state.page = "smart_lookup"
#             st.rerun()
#         st.markdown('</div>', unsafe_allow_html=True)

# # ---------------- MERGE FILES (UNCHANGED) ----------------
# elif st.session_state.page == "merge":

#     st.markdown('<div class="logo">📂 Merge Files</div>', unsafe_allow_html=True)

#     file_main = st.file_uploader("Upload Main File", type=["xlsx"])
#     file_status = st.file_uploader("Upload Status File", type=["xlsx"])
#     file_days = st.file_uploader("Upload Days File", type=["xlsx"])

#     if file_main and file_status and file_days:

#         df_main = pd.read_excel(file_main)
#         df_status = pd.read_excel(file_status)
#         df_days = pd.read_excel(file_days)

#         lot_col = None
#         for col in df_main.columns:
#             if "lot" in col.lower():
#                 lot_col = col
#                 break

#         if lot_col is None:
#             st.error("Lot column not found")
#             st.stop()

#         temp_file = "merge_temp.xlsx"

#         with pd.ExcelWriter(temp_file, engine="openpyxl") as writer:
#             df_main.to_excel(writer, sheet_name="Main", index=False)
#             df_status.to_excel(writer, sheet_name="Status", index=False)
#             df_days.to_excel(writer, sheet_name="Days", index=False)

#         wb = load_workbook(temp_file)
#         ws = wb["Main"]

#         lot_index = list(df_main.columns).index(lot_col) + 1
#         lot_letter = get_column_letter(lot_index)

#         status_col = ws.max_column + 1
#         ws.cell(row=1, column=status_col).value = "Status"

#         days_col = ws.max_column + 1
#         ws.cell(row=1, column=days_col).value = "No_of_Days"

#         for r in range(2, ws.max_row + 1):
#             ws.cell(row=r, column=status_col).value = f'=IFERROR(VLOOKUP({lot_letter}{r},Status!A:B,2,FALSE),"Not Found")'
#             ws.cell(row=r, column=days_col).value = f'=IFERROR(VLOOKUP({lot_letter}{r},Days!A:B,2,FALSE),"Not Found")'

#         output = io.BytesIO()
#         wb.save(output)
#         output.seek(0)

#         st.download_button("📥 Download Merged File", output, "diamond_merged.xlsx")

#     if st.button("⬅ Back"):
#         st.session_state.page = "home"
#         st.rerun()

# # ---------------- AUTOMATION (UNCHANGED) ----------------
# elif st.session_state.page == "automation":

#     st.markdown('<div class="logo">⚙️ Diamond Automation</div>', unsafe_allow_html=True)

#     uploaded_file = st.file_uploader("Upload Excel File", type=["xlsx"])

#     if uploaded_file:

#         df = pd.read_excel(uploaded_file)

#         lot_column = next((c for c in df.columns if "lot" in c.lower()), None)
#         if not lot_column:
#             st.error("Lot column not found")
#             st.stop()

#         cts_column = st.selectbox("Select CTS Column", df.columns)

#         if st.button("🚀 Run Automation"):

#             df[cts_column] = pd.to_numeric(df[cts_column], errors='coerce')

#             output = io.BytesIO()
#             df.to_excel(output, index=False)
#             output.seek(0)

#             st.download_button("📥 Download Excel File", output, "diamond_output.xlsx")

#     if st.button("⬅ Back"):
#         st.session_state.page = "home"
#         st.rerun()

# # ---------------- SMART MULTI LOOKUP (SAFE - FORMULA PRESERVED) ----------------
# elif st.session_state.page == "smart_lookup":

#     st.markdown('<div class="logo">⚡ Smart Multi Lookup</div>', unsafe_allow_html=True)

#     main_file = st.file_uploader("Upload Main File", type=["xlsx"])
#     lookup_files = st.file_uploader("Upload Team Files", type=["xlsx"], accept_multiple_files=True)

#     if main_file and lookup_files:

#         # Load main file as workbook (IMPORTANT)
#         wb = load_workbook(main_file)
#         ws = wb.active

#         df_main = pd.read_excel(main_file)

#         # 🔍 Detect LOT column
#         lot_col = None
#         for col in df_main.columns:
#             if "lot" in col.lower():
#                 lot_col = col
#                 break

#         if not lot_col:
#             st.error("Lot column not found")
#             st.stop()

#         # 🔍 Detect Updated Price column
#         updated_col = None
#         for col in df_main.columns:
#             if "updated" in col.lower() and "price" in col.lower():
#                 updated_col = col
#                 break

#         if not updated_col:
#             st.error("Updated Price column not found in main file")
#             st.stop()

#         lot_idx = list(df_main.columns).index(lot_col) + 1
#         upd_idx = list(df_main.columns).index(updated_col) + 1

#         lot_letter = get_column_letter(lot_idx)
#         upd_letter = get_column_letter(upd_idx)

#         # 📥 Combine lookup data
#         lookup_dict = {}

#         for file in lookup_files:

#             df_temp = pd.read_excel(file)

#             temp_lot = None
#             temp_price = None

#             for col in df_temp.columns:
#                 if "lot" in col.lower():
#                     temp_lot = col
#                 if "price" in col.lower():
#                     temp_price = col

#             if temp_lot and temp_price:
#                 for _, row in df_temp.iterrows():
#                     lookup_dict[row[temp_lot]] = row[temp_price]

#         if not lookup_dict:
#             st.error("No valid lookup data found")
#             st.stop()

#         # 🚀 APPLY VALUES (ONLY UPDATED PRICE COLUMN)
#         for r in range(2, ws.max_row + 1):

#             lot_value = ws.cell(row=r, column=lot_idx).value

#             current_value = ws.cell(row=r, column=upd_idx).value

#             if (current_value is None or str(current_value).strip() == "") and lot_value in lookup_dict:
#                  ws.cell(row=r, column=upd_idx).value = lookup_dict[lot_value]

#         st.success("✅ Updated Price filled without affecting formulas")

#         # 📥 Download
#         output = io.BytesIO()
#         wb.save(output)
#         output.seek(0)

#         st.download_button(
#             label="📥 Download Final File",
#             data=output,
#             file_name="final_output.xlsx"
#         )

#     col1, col2 = st.columns(2)

#     with col1:
#         if st.button("⬅ Back"):
#             st.session_state.page = "home"
#             st.rerun()

#     with col2:
#         if st.button("Logout"):
#             st.session_state.login = False
#             st.session_state.page = "login"
#             st.rerun()



# import streamlit as st

# # ---------------- CONFIG ----------------
# st.set_page_config(page_title="GEMAI", page_icon="💎", layout="wide")

# # ---------------- REMOVE SIDEBAR ----------------
# st.markdown("""
# <style>
# [data-testid="stSidebar"] {display: none;}
# </style>
# """, unsafe_allow_html=True)

# # ---------------- STYLE ----------------
# st.markdown("""
# <style>
# .stApp {background: linear-gradient(135deg, #0f2027, #000000); color: white;}
# .logo {font-size: 30px; font-weight: bold; color: #00FFD1;}
# .card {padding: 25px; border-radius: 12px; background: #111; border: 2px solid #1f2a30;}
# .stButton>button {background: #00FFD1; color: black; border-radius: 8px; font-weight: bold;}
# .login-box {max-width: 400px; margin: auto; margin-top: 120px; padding: 30px; background: #111; border-radius: 12px; border: 1px solid #00FFD1;}
# </style>
# """, unsafe_allow_html=True)

# # ---------------- SESSION ----------------
# if "login" not in st.session_state:
#     st.session_state.login = False
# if "page" not in st.session_state:
#     st.session_state.page = "login"

# # ---------------- LOGIN ----------------
# if not st.session_state.login:

#     col1, col2, col3 = st.columns([2,3,2])
#     with col2:
#         st.markdown('<div class="login-box">', unsafe_allow_html=True)
#         st.markdown('<div class="logo">💎 GEMAI</div>', unsafe_allow_html=True)

#         username = st.text_input("Username")
#         password = st.text_input("Password", type="password")

#         if st.button("Login", use_container_width=True):
#             if username == "admin" and password == "1234":
#                 st.session_state.login = True
#                 st.session_state.page = "home"
#                 st.rerun()
#             else:
#                 st.error("Invalid Credentials")

#         st.markdown('</div>', unsafe_allow_html=True)

# # ---------------- HOME ----------------
# elif st.session_state.page == "home":

#     col1, col2 = st.columns([6,1])

#     with col1:
#         st.markdown('<div class="logo">💎 GEMAI</div>', unsafe_allow_html=True)

#     with col2:
#         if st.button("Logout"):
#             st.session_state.login = False
#             st.session_state.page = "login"
#             st.rerun()

#     st.subheader("Dashboard")

#     col1, col2, col3 = st.columns(3)

#     if st.button("📂 Merge Files"):
#         st.session_state.page = "merge"
#         st.rerun()

#     if st.button("⚙️ Automation"):
#         st.session_state.page = "automation"
#         st.rerun()

#     if st.button("⚡ Smart Lookup"):
#         st.session_state.page = "smart"
#         st.rerun()

# # ---------------- ROUTING ----------------
# elif st.session_state.page == "merge":
#     from merge import show_merge
#     show_merge()

# elif st.session_state.page == "automation":
#     from automation import show_automation
#     show_automation()

# elif st.session_state.page == "smart":
#     from smart_lookup import show_smart_lookup
#     show_smart_lookup()


# import streamlit as st
# import pandas as pd
# from openpyxl import load_workbook
# from openpyxl.utils import get_column_letter
# import io
# from automation import show_automation  # This is CORRECT in app.py

# # ---------------- CONFIG ----------------
# st.set_page_config(page_title="GEMAI", page_icon="💎", layout="wide")

# # ---------------- REMOVE SIDEBAR ----------------
# st.markdown("""
# <style>
# [data-testid="stSidebar"] {display: none;}
# </style>
# """, unsafe_allow_html=True)

# # ---------------- UI STYLE ----------------
# st.markdown("""
# <style>
# .stApp {
#     background: linear-gradient(135deg, #0f2027, #000000);
#     color: white;
# }
# .logo {
#     font-size: 30px;
#     font-weight: bold;
#     color: #00FFD1;
# }
# .card {
#     padding: 25px;
#     border-radius: 12px;
#     background: #111;
#     border: 2px solid #1f2a30;
#     transition: 0.3s;
# }
# .card:hover {
#     border: 2px solid #00FFD1;
#     box-shadow: 0px 0px 20px #00FFD1;
# }
# .stButton>button {
#     background: #00FFD1;
#     color: black;
#     border-radius: 8px;
#     font-weight: bold;
# }
# .login-box {
#     max-width: 400px;
#     margin: auto;
#     margin-top: 120px;
#     padding: 30px;
#     background: #111;
#     border-radius: 12px;
#     border: 1px solid #00FFD1;
# }
# </style>
# """, unsafe_allow_html=True)

# # ---------------- SESSION ----------------
# if "login" not in st.session_state:
#     st.session_state.login = False
# if "page" not in st.session_state:
#     st.session_state.page = "login"

# # ---------------- LOGIN ----------------
# if not st.session_state.login:
#     col1, col2, col3 = st.columns([2, 3, 2])
#     with col2:
#         st.markdown('<div class="login-box">', unsafe_allow_html=True)
#         st.markdown('<div class="logo">💎 GEMAI</div>', unsafe_allow_html=True)
#         username = st.text_input("Username")
#         password = st.text_input("Password", type="password")
#         if st.button("Login", use_container_width=True):
#             if username == "admin" and password == "1234":
#                 st.session_state.login = True
#                 st.session_state.page = "home"
#                 st.rerun()
#             else:
#                 st.error("Invalid Credentials")
#         st.markdown('</div>', unsafe_allow_html=True)

# # ---------------- HOME ----------------
# elif st.session_state.page == "home":
#     col1, col2 = st.columns([6, 1])
#     with col1:
#         st.markdown('<div class="logo">💎 GEMAI</div>', unsafe_allow_html=True)
#     with col2:
#         if st.button("Logout"):
#             st.session_state.login = False
#             st.session_state.page = "login"
#             st.rerun()

#     st.write("")
#     st.subheader("Dashboard")

#     col1, col2 = st.columns(2)
#     with col1:
#         st.markdown('<div class="card">', unsafe_allow_html=True)
#         st.subheader("📂 Merge + Automate")
#         st.caption("Upload files, merge status/days if provided, then run full automation.")
#         if st.button("Open Tool"):
#             st.session_state.page = "unified"
#             st.rerun()
#         st.markdown('</div>', unsafe_allow_html=True)

#     with col2:
#         st.markdown('<div class="card">', unsafe_allow_html=True)
#         st.subheader("ℹ️ How It Works")
#         st.caption(
#             "1. Upload Main file (required)\n"
#             "2. Upload Status & Days files (optional — for merging)\n"
#             "3. Select CTS column\n"
#             "4. Run → Download single output file"
#         )
#         st.markdown('</div>', unsafe_allow_html=True)



# # ---------------- UNIFIED PAGE ----------------
# elif st.session_state.page == "unified":

#     col1, col2 = st.columns([6, 1])
#     with col1:
#         st.markdown('<div class="logo">💎 GEMAI — Merge & Automate</div>', unsafe_allow_html=True)
#     with col2:
#         if st.button("Logout"):
#             st.session_state.login = False
#             st.session_state.page = "login"
#             st.rerun()

#     st.write("")

#     # ---- FILE UPLOADS ----
#     st.subheader("📁 Upload Files")
#     file_main = st.file_uploader("Upload Main File (required)", type=["xlsx"])

#     col_s, col_d = st.columns(2)
#     with col_s:
#         file_status = st.file_uploader("Upload Status File (optional)", type=["xlsx"])
#     with col_d:
#         file_days = st.file_uploader("Upload Days File (optional)", type=["xlsx"])

#     if file_main:
#         df_main = pd.read_excel(file_main)

#         # ---- DETECT LOT COLUMN ----
#         lot_col = next((col for col in df_main.columns if "lot" in col.lower()), None)
#         if lot_col is None:
#             st.error("❌ Lot column not found in Main file.")
#             st.stop()

#         st.success(f"✅ Lot column detected: **{lot_col}**")

#         # ---- MERGE STEP (if status + days provided) ----
#         do_merge = file_status is not None and file_days is not None

#         if file_status and not file_days:
#             st.warning("⚠️ Status file uploaded but Days file is missing — skipping merge.")
#         if file_days and not file_status:
#             st.warning("⚠️ Days file uploaded but Status file is missing — skipping merge.")

#         if do_merge:
#             st.info("🔗 Status & Days files detected — merge will be applied before automation.")

#         # ---- SELECT CTS COLUMN ----
#         st.subheader("⚙️ Automation Settings")
#         cts_column = st.selectbox("Select CTS Column", df_main.columns)

#         st.dataframe(df_main.head(), use_container_width=True)

#         # ---- RUN BUTTON ----
#         if st.button("🚀 Run"):
#             try:
#                 # Call the automation function
#                 output = show_automation(
#                     df_main=df_main,
#                     file_main=file_main,
#                     file_status=file_status,
#                     file_days=file_days,
#                     lot_col=lot_col,
#                     cts_column=cts_column,
#                     do_merge=do_merge
#                 )
                
#                 label = "📥 Download Merged & Automated File" if do_merge else "📥 Download Automated File"
#                 filename = "diamond_merged_automated.xlsx" if do_merge else "diamond_automated.xlsx"

#                 st.success("✅ Done! Click below to download your file.")
#                 st.download_button(label=label, data=output, file_name=filename)
                
#             except Exception as e:
#                 st.error(f"❌ Error during automation: {str(e)}")

#     st.write("")
#     if st.button("⬅ Back to Dashboard"):
#         st.session_state.page = "home"
#         st.rerun()

import streamlit as st
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import io
from automation import show_automation

# ---------------- CONFIG ----------------
st.set_page_config(page_title="GEMAI", page_icon="💎", layout="wide")

# ---------------- REMOVE SIDEBAR ----------------
st.markdown("""
<style>
[data-testid="stSidebar"] {display: none;}
</style>
""", unsafe_allow_html=True)

# ---------------- UI STYLE ----------------
st.markdown("""
<style>
.stApp {
    background: linear-gradient(135deg, #0f2027, #000000);
    color: white;
}
.logo {
    font-size: 30px;
    font-weight: bold;
    color: #00FFD1;
}
.card {
    padding: 25px;
    border-radius: 12px;
    background: #111;
    border: 2px solid #1f2a30;
    transition: 0.3s;
    margin-bottom: 20px;
}
.card:hover {
    border: 2px solid #00FFD1;
    box-shadow: 0px 0px 20px #00FFD1;
}
.stButton>button {
    background: #00FFD1;
    color: black;
    border-radius: 8px;
    font-weight: bold;
}
.login-box {
    max-width: 400px;
    margin: auto;
    margin-top: 120px;
    padding: 30px;
    background: #111;
    border-radius: 12px;
    border: 1px solid #00FFD1;
}
</style>
""", unsafe_allow_html=True)

# ---------------- SESSION ----------------
if "login" not in st.session_state:
    st.session_state.login = False
if "page" not in st.session_state:
    st.session_state.page = "login"

# ---------------- SMART LOOKUP FUNCTION (defined here for access) ----------------
def show_smart_lookup():
    st.markdown('<div class="logo">⚡ Smart Lookup</div>', unsafe_allow_html=True)
    
    # Add option selection
    lookup_option = st.radio(
        "Select Lookup Mode:",
        ["Fill Empty Updated Price", "Update All Prices (Overwrite)"],
        help="Fill Empty: Only updates blank cells | Update All: Overwrites all prices from team files",
        horizontal=True
    )
    
    main_file = st.file_uploader("Upload Main File", type=["xlsx"])
    files = st.file_uploader("Upload Team Files", type=["xlsx"], accept_multiple_files=True)

    if main_file and files:
        try:
            wb = load_workbook(main_file)
            ws = wb.active

            df = pd.read_excel(main_file)

            lot_col = next((c for c in df.columns if "lot" in c.lower()), None)
            upd_col = next((c for c in df.columns if "updated" in c.lower()), None)

            if lot_col is None:
                st.error("❌ Lot column not found in Main file")
                return
                
            if upd_col is None:
                st.error("❌ Updated Price column not found in Main file")
                return

            lot_idx = list(df.columns).index(lot_col) + 1
            upd_idx = list(df.columns).index(upd_col) + 1

            lookup = {}
            files_processed = 0

            for f in files:
                temp = pd.read_excel(f)
                l = next((c for c in temp.columns if "lot" in c.lower()), None)
                p = next((c for c in temp.columns if "price" in c.lower()), None)

                if l and p:
                    for _, row in temp.iterrows():
                        lookup[row[l]] = row[p]
                    files_processed += 1
                else:
                    st.warning(f"⚠️ File {f.name} missing Lot or Price column - skipping")

            if files_processed == 0:
                st.error("❌ No valid team files found with Lot and Price columns")
                return

            st.info(f"📊 Loaded {len(lookup)} unique lot prices from {files_processed} file(s)")

            # Counter for updated cells
            updated_count = 0
            skipped_count = 0
            
            for r in range(2, ws.max_row+1):
                lot_val = ws.cell(r, lot_idx).value
                current = ws.cell(r, upd_idx).value
                
                # Check if lot exists in lookup
                if lot_val in lookup:
                    if lookup_option == "Fill Empty Updated Price":
                        # Original logic: only fill if empty
                        if current is None or str(current).strip() == "":
                            ws.cell(r, upd_idx).value = lookup[lot_val]
                            updated_count += 1
                        else:
                            skipped_count += 1
                    else:  # "Update All Prices (Overwrite)"
                        # New logic: update all prices regardless of existing value
                        ws.cell(r, upd_idx).value = lookup[lot_val]
                        updated_count += 1

            output = io.BytesIO()
            wb.save(output)
            output.seek(0)

            # Show summary
            st.success(f"✅ Updated {updated_count} cells using '{lookup_option}' mode")
            if lookup_option == "Fill Empty Updated Price" and skipped_count > 0:
                st.info(f"ℹ️ Skipped {skipped_count} cells that already had values")
            
            st.download_button("📥 Download Final File", output, "final_output.xlsx", use_container_width=True)
            
        except Exception as e:
            st.error(f"❌ Error: {str(e)}")

    if st.button("⬅ Back to Dashboard", use_container_width=True):
        st.session_state.page = "home"
        st.rerun()

# ---------------- LOGIN ----------------
if not st.session_state.login:
    col1, col2, col3 = st.columns([2, 3, 2])
    with col2:
        st.markdown('<div class="login-box">', unsafe_allow_html=True)
        st.markdown('<div class="logo">💎 GEMAI</div>', unsafe_allow_html=True)
        username = st.text_input("Username")
        password = st.text_input("Password", type="password")
        if st.button("Login", use_container_width=True):
            if username == "admin" and password == "1234":
                st.session_state.login = True
                st.session_state.page = "home"
                st.rerun()
            else:
                st.error("Invalid Credentials")
        st.markdown('</div>', unsafe_allow_html=True)

# ---------------- HOME ----------------
elif st.session_state.page == "home":
    col1, col2 = st.columns([6, 1])
    with col1:
        st.markdown('<div class="logo">💎 GEMAI</div>', unsafe_allow_html=True)
    with col2:
        if st.button("Logout"):
            st.session_state.login = False
            st.session_state.page = "login"
            st.rerun()

    st.write("")
    st.subheader("Dashboard")
    st.write("")

    # Create 3 columns for cards
    col1, col2, col3 = st.columns(3)
    
    with col1:
        st.markdown('<div class="card">', unsafe_allow_html=True)
        st.subheader("📂 Merge + Automate")
        st.caption("Upload files, merge status/days if provided, then run full automation with price validation.")
        if st.button("Open Tool", key="open_automate", use_container_width=True):
            st.session_state.page = "unified"
            st.rerun()
        st.markdown('</div>', unsafe_allow_html=True)

    with col2:
        st.markdown('<div class="card">', unsafe_allow_html=True)
        st.subheader("⚡ Smart Lookup")
        st.caption("Fill missing Updated Prices from team files using Lot number matching. Two modes available.")
        if st.button("Open Smart Lookup", key="open_lookup", use_container_width=True):
            st.session_state.page = "smart_lookup"
            st.rerun()
        st.markdown('</div>', unsafe_allow_html=True)

    with col3:
        st.markdown('<div class="card">', unsafe_allow_html=True)
        st.subheader("ℹ️ Features")
        st.caption(
            "✓ Merge Status & Days files\n"
            "✓ Auto-detect Lot & CTS columns\n"
            "✓ Size Group classification\n"
            "✓ Price validation with highlighting\n"
            "✓ Smart Lookup from team files\n"
            "✓ U-series & <1ct formulas"
        )
        st.markdown('</div>', unsafe_allow_html=True)

# ---------------- UNIFIED PAGE (Merge + Automate) ----------------
elif st.session_state.page == "unified":

    col1, col2 = st.columns([6, 1])
    with col1:
        st.markdown('<div class="logo">💎 GEMAI — Merge & Automate</div>', unsafe_allow_html=True)
    with col2:
        if st.button("Logout"):
            st.session_state.login = False
            st.session_state.page = "login"
            st.rerun()

    st.write("")

    # ---- FILE UPLOADS ----
    st.subheader("📁 Upload Files")
    file_main = st.file_uploader("Upload Main File (required)", type=["xlsx"])

    col_s, col_d = st.columns(2)
    with col_s:
        file_status = st.file_uploader("Upload Status File (optional)", type=["xlsx"])
    with col_d:
        file_days = st.file_uploader("Upload Days File (optional)", type=["xlsx"])

    if file_main:
        df_main = pd.read_excel(file_main)

        # ---- DETECT LOT COLUMN ----
        lot_col = next((col for col in df_main.columns if "lot" in col.lower()), None)
        if lot_col is None:
            st.error("❌ Lot column not found in Main file.")
            st.stop()

        st.success(f"✅ Lot column detected: **{lot_col}**")

        # ---- MERGE STEP (if status + days provided) ----
        do_merge = file_status is not None and file_days is not None

        if file_status and not file_days:
            st.warning("⚠️ Status file uploaded but Days file is missing — skipping merge.")
        if file_days and not file_status:
            st.warning("⚠️ Days file uploaded but Status file is missing — skipping merge.")

        if do_merge:
            st.info("🔗 Status & Days files detected — merge will be applied before automation.")

        # ---- SELECT CTS COLUMN ----
        st.subheader("⚙️ Automation Settings")
        cts_column = st.selectbox("Select CTS Column", df_main.columns)

        st.dataframe(df_main.head(), use_container_width=True)

        # ---- RUN BUTTON ----
        if st.button("🚀 Run", use_container_width=True):
            try:
                # Call the automation function
                output = show_automation(
                    df_main=df_main,
                    file_main=file_main,
                    file_status=file_status,
                    file_days=file_days,
                    lot_col=lot_col,
                    cts_column=cts_column,
                    do_merge=do_merge
                )
                
                label = "📥 Download Merged & Automated File" if do_merge else "📥 Download Automated File"
                filename = "diamond_merged_automated.xlsx" if do_merge else "diamond_automated.xlsx"

                st.success("✅ Done! Click below to download your file.")
                st.download_button(label=label, data=output, file_name=filename, use_container_width=True)
                
            except Exception as e:
                st.error(f"❌ Error during automation: {str(e)}")

    st.write("")
    if st.button("⬅ Back to Dashboard", use_container_width=True):
        st.session_state.page = "home"
        st.rerun()

# ---------------- SMART LOOKUP PAGE ----------------
elif st.session_state.page == "smart_lookup":
    show_smart_lookup()