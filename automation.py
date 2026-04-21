# import pandas as pd
# from openpyxl import load_workbook
# from openpyxl.utils import get_column_letter
# import io

# def show_automation(df_main, file_main, file_status, file_days, lot_col, cts_column, do_merge):
#     """
#     Main automation function that handles merging and automation logic
#     """
    
#     # ── STEP 1: Merge if needed ──
#     if do_merge:
#         df_status = pd.read_excel(file_status)
#         df_days = pd.read_excel(file_days)

#         temp_merge = "merge_temp.xlsx"
#         with pd.ExcelWriter(temp_merge, engine="openpyxl") as writer:
#             df_main.to_excel(writer, sheet_name="Main", index=False)
#             df_status.to_excel(writer, sheet_name="Status", index=False)
#             df_days.to_excel(writer, sheet_name="Days", index=False)

#         wb_merge = load_workbook(temp_merge)
#         ws_merge = wb_merge["Main"]

#         lot_index = list(df_main.columns).index(lot_col) + 1
#         lot_letter = get_column_letter(lot_index)

#         status_col_idx = ws_merge.max_column + 1
#         ws_merge.cell(row=1, column=status_col_idx).value = "Status"

#         days_col_idx = ws_merge.max_column + 1
#         ws_merge.cell(row=1, column=days_col_idx).value = "No_of_Days"

#         for r in range(2, ws_merge.max_row + 1):
#             ws_merge.cell(row=r, column=status_col_idx).value = (
#                 f'=IFERROR(VLOOKUP({lot_letter}{r},Status!A:B,2,FALSE),"Not Found")'
#             )
#             ws_merge.cell(row=r, column=days_col_idx).value = (
#                 f'=IFERROR(VLOOKUP({lot_letter}{r},Days!A:B,2,FALSE),"Not Found")'
#             )

#         # Save merged workbook to bytes and reload as DataFrame for automation
#         merged_bytes = io.BytesIO()
#         wb_merge.save(merged_bytes)
#         merged_bytes.seek(0)

#         # Read back Main sheet values only (formulas stay in xlsx, df gets current values)
#         # We keep the openpyxl workbook for automation to preserve formulas
#         wb = wb_merge
#         ws = wb["Main"]

#         # Rebuild df column list from merged sheet header row
#         df_cols = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]

#     else:
#         # No merge — load fresh workbook from main file
#         file_main.seek(0)
#         temp_auto = "auto_temp.xlsx"
#         df_main.to_excel(temp_auto, index=False)
#         wb = load_workbook(temp_auto)
#         ws = wb.active
#         df_cols = list(df_main.columns)

#     # ── STEP 2: Automation ──
#     # Find CTS column index in workbook
#     try:
#         col_index = df_cols.index(cts_column) + 1
#     except ValueError:
#         raise ValueError(f"CTS column '{cts_column}' not found in workbook headers.")

#     # Find cost column index (Price/CTS or Cost/CTS)
#     cost_col = None
#     for i, col in enumerate(df_cols):
#         if col and ("price" in str(col).lower() or "cost" in str(col).lower()):
#             cost_col = i + 1
#             break
    
#     if cost_col is None:
#         raise ValueError("Price/Cost column not found in Main file.")

#     # Find Lot column index
#     lot_index = None
#     for i, col in enumerate(df_cols):
#         if col and "lot" in str(col).lower():
#             lot_index = i + 1
#             break
    
#     if lot_index is None:
#         raise ValueError("Lot column not found in Main file.")

#     # Convert CTS column to numeric in df_main for value checks
#     df_main[cts_column] = pd.to_numeric(df_main[cts_column], errors="coerce")

#     # SIZE MAP sheet
#     if "SizeMap" in wb.sheetnames:
#         del wb["SizeMap"]
#     ws2 = wb.create_sheet("SizeMap")

#     size_data = [
#         (0.30, "0.30-0.49"), (0.50, "0.50-0.59"), (0.60, "0.60-0.69"),
#         (0.70, "0.70-0.79"), (0.80, "0.80-0.89"), (0.90, "0.90-0.99"),
#         (1.00, "1.00-1.10"), (1.11, "1.11-1.49"), (1.50, "1.50-1.59"),
#         (1.60, "1.60-1.99"), (2.00, "2.00-2.10"), (2.11, "2.11-2.49"),
#         (2.50, "2.50-2.59"), (2.60, "2.60-2.99"), (3.00, "3.00-3.10"),
#         (3.11, "3.11-3.49"), (3.50, "3.50-3.59"), (3.60, "3.60-3.99"),
#         (4.00, "4.00-4.10"), (4.11, "4.11-4.49"),
#         (4.50, "4.50-4.59"), (4.60, "4.60-4.99"),
#         (5.00, "5.00-5.49"), (5.50, "5.50-5.99"),
#         (6.00, "6.00-6.99"), (7.00, "7.00-7.99"),
#         (8.00, "8.00-8.99"), (9.00, "9.00-9.99"),
#         (10.00, "10.00-10.99"), (11.00, "11.00-11.99"),
#         (12.00, "12.00-12.99"), (13.00, "13.00-13.99"),
#         (14.00, "14.00-14.99"), (15.00, "15.00-15.99"),
#         (16.00, "16.00-16.99"), (17.00, "17.00-17.99"),
#         (18.00, "18.00-18.99"), (19.00, "19.00-19.99"),
#         (20.00, "20.00-20.99"), (21.00, "21.00-21.99"),
#         (22.00, "22.00-22.99"), (23.00, "23.00-23.99"),
#         (24.00, "24.00-24.99"), (25.00, "25.00-25.99"),
#         (25.01, "High Carat Stone"),
#     ]
#     ws2.append(["Min CTS", "Size Group"])
#     for row in size_data:
#         ws2.append(row)

#     # Size Group column
#     size_col = ws.max_column + 1
#     ws.cell(row=1, column=size_col).value = "Size Group"

#     for r in range(2, ws.max_row + 1):
#         ws.cell(row=r, column=size_col).value = (
#             f'=VLOOKUP({get_column_letter(col_index)}{r},SizeMap!A:B,2,TRUE)'
#         )

#     # Updated Price column
#     up_col = ws.max_column + 1
#     ws.cell(row=1, column=up_col).value = "Updated Price"

#     # Difference (%) column
#     diff_col = ws.max_column + 1
#     ws.cell(row=1, column=diff_col).value = "Difference (%)"

#     for r in range(2, ws.max_row + 1):
#         c = f"{get_column_letter(cost_col)}{r}"
#         u = f"{get_column_letter(up_col)}{r}"
#         ws.cell(row=r, column=diff_col).value = f"=-ROUND(({c}-{u})/{c}%,2)"

#     # Cost Value / Updated Value / Profit Loss columns
#     # THESE FORMULAS ONLY APPLY TO U SERIES STONES AND CTS < 1
#     cv = ws.max_column + 1
#     uv = ws.max_column + 2
#     pf = ws.max_column + 3

#     ws.cell(row=1, column=cv).value = "Cost Value"
#     ws.cell(row=1, column=uv).value = "Updated Value"
#     ws.cell(row=1, column=pf).value = "Profit/Loss"

#     # Get lot column values to check for U series
#     lot_letter = get_column_letter(lot_index)
#     cts_letter = get_column_letter(col_index)
#     cost_letter = get_column_letter(cost_col)
#     upd_letter = get_column_letter(up_col)
    
#     # Apply formulas ONLY for U series stones OR CTS < 1
#     for r in range(2, ws.max_row + 1):
#         # Get lot value from cell
#         lot_cell = ws.cell(row=r, column=lot_index).value
#         lot_value = str(lot_cell) if lot_cell is not None else ""
        
#         # Get CTS value
#         cts_cell = ws.cell(row=r, column=col_index).value
#         try:
#             cts_value = float(cts_cell) if cts_cell is not None else None
#         except (ValueError, TypeError):
#             cts_value = None
        
#         # Check if this row should have formulas
#         is_u_series = lot_value.startswith('U') if lot_value else False
#         is_less_than_1 = cts_value is not None and cts_value < 1
        
#         if is_u_series or is_less_than_1:
#             # Apply formulas for Cost Value, Updated Value, Profit/Loss
#             ws.cell(row=r, column=cv).value = f"={cost_letter}{r}*{cts_letter}{r}"
#             ws.cell(row=r, column=uv).value = f"={upd_letter}{r}*{cts_letter}{r}"
#             ws.cell(row=r, column=pf).value = f"={get_column_letter(uv)}{r}-{get_column_letter(cv)}{r}"
#         else:
#             # Leave cells empty for other stones
#             ws.cell(row=r, column=cv).value = ""
#             ws.cell(row=r, column=uv).value = ""
#             ws.cell(row=r, column=pf).value = ""

#     # ── STEP 3: Save & Return ──
#     output = io.BytesIO()
#     wb.save(output)
#     output.seek(0)
    
#     return output


import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font
import io

def show_automation(df_main, file_main, file_status, file_days, lot_col, cts_column, do_merge):
    """
    Main automation function that handles merging and automation logic
    """
    
    # Define color ranking (higher number = better quality)
    color_ranking = {
        'D': 10, 'E': 9, 'F': 8, 'G': 7, 'H': 6, 'I': 5, 'J': 4, 'K': 3, 'L': 2, 'M': 1
    }
    
    # Define clarity ranking (higher number = better quality)
    clarity_ranking = {
        'IF': 12, 'VVS1': 11, 'VVS2': 10, 'VS1': 9, 'VS2': 8, 
        'SI1': 7, 'SI2': 6, 'I1': 5, 'I2': 4, 'I3': 3
    }
    
    # ── STEP 1: Merge if needed ──
    if do_merge:
        df_status = pd.read_excel(file_status)
        df_days = pd.read_excel(file_days)

        temp_merge = "merge_temp.xlsx"
        with pd.ExcelWriter(temp_merge, engine="openpyxl") as writer:
            df_main.to_excel(writer, sheet_name="Main", index=False)
            df_status.to_excel(writer, sheet_name="Status", index=False)
            df_days.to_excel(writer, sheet_name="Days", index=False)

        wb_merge = load_workbook(temp_merge)
        ws_merge = wb_merge["Main"]

        lot_index = list(df_main.columns).index(lot_col) + 1
        lot_letter = get_column_letter(lot_index)

        status_col_idx = ws_merge.max_column + 1
        ws_merge.cell(row=1, column=status_col_idx).value = "Status"

        days_col_idx = ws_merge.max_column + 1
        ws_merge.cell(row=1, column=days_col_idx).value = "No_of_Days"

        for r in range(2, ws_merge.max_row + 1):
            ws_merge.cell(row=r, column=status_col_idx).value = (
                f'=IFERROR(VLOOKUP({lot_letter}{r},Status!A:B,2,FALSE),"Not Found")'
            )
            ws_merge.cell(row=r, column=days_col_idx).value = (
                f'=IFERROR(VLOOKUP({lot_letter}{r},Days!A:B,2,FALSE),"Not Found")'
            )

        # Save merged workbook to bytes and reload as DataFrame for automation
        merged_bytes = io.BytesIO()
        wb_merge.save(merged_bytes)
        merged_bytes.seek(0)

        # Read back Main sheet values only (formulas stay in xlsx, df gets current values)
        wb = wb_merge
        ws = wb["Main"]

        # Rebuild df column list from merged sheet header row
        df_cols = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]

    else:
        # No merge — load fresh workbook from main file
        file_main.seek(0)
        temp_auto = "auto_temp.xlsx"
        df_main.to_excel(temp_auto, index=False)
        wb = load_workbook(temp_auto)
        ws = wb.active
        df_cols = list(df_main.columns)

    # ── STEP 2: Automation ──
    # Find CTS column index in workbook
    try:
        col_index = df_cols.index(cts_column) + 1
    except ValueError:
        raise ValueError(f"CTS column '{cts_column}' not found in workbook headers.")

    # Find cost column index (Price/CTS or Cost/CTS)
    cost_col = None
    for i, col in enumerate(df_cols):
        if col and ("price" in str(col).lower() or "cost" in str(col).lower()):
            cost_col = i + 1
            break
    
    if cost_col is None:
        raise ValueError("Price/Cost column not found in Main file.")

    # Find Lot column index
    lot_index = None
    for i, col in enumerate(df_cols):
        if col and "lot" in str(col).lower():
            lot_index = i + 1
            break
    
    if lot_index is None:
        raise ValueError("Lot column not found in Main file.")
    
    # Find Color column index
    color_index = None
    for i, col in enumerate(df_cols):
        if col and "color" in str(col).lower():
            color_index = i + 1
            break
    
    # Find Clarity column index
    clarity_index = None
    for i, col in enumerate(df_cols):
        if col and "clarity" in str(col).lower():
            clarity_index = i + 1
            break

    # Convert CTS column to numeric in df_main for value checks
    df_main[cts_column] = pd.to_numeric(df_main[cts_column], errors="coerce")

    # SIZE MAP sheet
    if "SizeMap" in wb.sheetnames:
        del wb["SizeMap"]
    ws2 = wb.create_sheet("SizeMap")

    size_data = [
        (0.30, "0.30-0.49"), (0.50, "0.50-0.59"), (0.60, "0.60-0.69"),
        (0.70, "0.70-0.79"), (0.80, "0.80-0.89"), (0.90, "0.90-0.99"),
        (1.00, "1.00-1.10"), (1.11, "1.11-1.49"), (1.50, "1.50-1.59"),
        (1.60, "1.60-1.99"), (2.00, "2.00-2.10"), (2.11, "2.11-2.49"),
        (2.50, "2.50-2.59"), (2.60, "2.60-2.99"), (3.00, "3.00-3.10"),
        (3.11, "3.11-3.49"), (3.50, "3.50-3.59"), (3.60, "3.60-3.99"),
        (4.00, "4.00-4.10"), (4.11, "4.11-4.49"),
        (4.50, "4.50-4.59"), (4.60, "4.60-4.99"),
        (5.00, "5.00-5.49"), (5.50, "5.50-5.99"),
        (6.00, "6.00-6.99"), (7.00, "7.00-7.99"),
        (8.00, "8.00-8.99"), (9.00, "9.00-9.99"),
        (10.00, "10.00-10.99"), (11.00, "11.00-11.99"),
        (12.00, "12.00-12.99"), (13.00, "13.00-13.99"),
        (14.00, "14.00-14.99"), (15.00, "15.00-15.99"),
        (16.00, "16.00-16.99"), (17.00, "17.00-17.99"),
        (18.00, "18.00-18.99"), (19.00, "19.00-19.99"),
        (20.00, "20.00-20.99"), (21.00, "21.00-21.99"),
        (22.00, "22.00-22.99"), (23.00, "23.00-23.99"),
        (24.00, "24.00-24.99"), (25.00, "25.00-25.99"),
        (25.01, "High Carat Stone"),
    ]
    ws2.append(["Min CTS", "Size Group"])
    for row in size_data:
        ws2.append(row)

    # Size Group column
    size_col = ws.max_column + 1
    ws.cell(row=1, column=size_col).value = "Size Group"

    for r in range(2, ws.max_row + 1):
        ws.cell(row=r, column=size_col).value = (
            f'=VLOOKUP({get_column_letter(col_index)}{r},SizeMap!A:B,2,TRUE)'
        )

    # Updated Price column
    up_col = ws.max_column + 1
    ws.cell(row=1, column=up_col).value = "Updated Price"

    # Difference (%) column
    diff_col = ws.max_column + 1
    ws.cell(row=1, column=diff_col).value = "Difference (%)"

    for r in range(2, ws.max_row + 1):
        c = f"{get_column_letter(cost_col)}{r}"
        u = f"{get_column_letter(up_col)}{r}"
        ws.cell(row=r, column=diff_col).value = f"=-ROUND(({c}-{u})/{c}%,2)"

    # Cost Value / Updated Value / Profit Loss columns
    cv = ws.max_column + 1
    uv = ws.max_column + 2
    pf = ws.max_column + 3

    ws.cell(row=1, column=cv).value = "Cost Value"
    ws.cell(row=1, column=uv).value = "Updated Value"
    ws.cell(row=1, column=pf).value = "Profit/Loss"

    # Get column letters
    lot_letter = get_column_letter(lot_index)
    cts_letter = get_column_letter(col_index)
    cost_letter = get_column_letter(cost_col)
    upd_letter = get_column_letter(up_col)
    
    # First pass: Apply formulas for U series or CTS < 1
    for r in range(2, ws.max_row + 1):
        lot_cell = ws.cell(row=r, column=lot_index).value
        lot_value = str(lot_cell) if lot_cell is not None else ""
        
        cts_cell = ws.cell(row=r, column=col_index).value
        try:
            cts_value = float(cts_cell) if cts_cell is not None else None
        except (ValueError, TypeError):
            cts_value = None
        
        is_u_series = lot_value.startswith('U') if lot_value else False
        is_less_than_1 = cts_value is not None and cts_value < 1
        
        if is_u_series or is_less_than_1:
            ws.cell(row=r, column=cv).value = f"={cost_letter}{r}*{cts_letter}{r}"
            ws.cell(row=r, column=uv).value = f"={upd_letter}{r}*{cts_letter}{r}"
            ws.cell(row=r, column=pf).value = f"={get_column_letter(uv)}{r}-{get_column_letter(cv)}{r}"

    # ── STEP 3: Price Validation & Highlighting ──
    if color_index and clarity_index:
        # Create a validation sheet for price hierarchy
        if "PriceValidation" in wb.sheetnames:
            del wb["PriceValidation"]
        ws_validation = wb.create_sheet("PriceValidation")
        
        # Add headers
        ws_validation.cell(row=1, column=1).value = "Size Group"
        ws_validation.cell(row=1, column=2).value = "Row"
        ws_validation.cell(row=1, column=3).value = "Color"
        ws_validation.cell(row=1, column=4).value = "Clarity"
        ws_validation.cell(row=1, column=5).value = "Quality Score"
        ws_validation.cell(row=1, column=6).value = "Updated Price"
        ws_validation.cell(row=1, column=7).value = "Expected Max Price"
        ws_validation.cell(row=1, column=8).value = "Status"
        
        # Create fills
        red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        
        # Collect data for each size group
        size_groups = {}
        
        # Read all data first
        for r in range(2, ws.max_row + 1):
            size_group = ws.cell(row=r, column=size_col).value
            if size_group:
                if size_group not in size_groups:
                    size_groups[size_group] = []
                
                color = ws.cell(row=r, column=color_index).value if color_index else None
                clarity = ws.cell(row=r, column=clarity_index).value if clarity_index else None
                updated_price = ws.cell(row=r, column=up_col).value
                
                # Get quality scores
                color_str = str(color).strip().upper() if color else ''
                clarity_str = str(clarity).strip().upper() if clarity else ''
                
                color_score = color_ranking.get(color_str, 0)
                clarity_score = clarity_ranking.get(clarity_str, 0)
                total_quality_score = color_score + clarity_score
                
                # Try to convert price to float
                try:
                    price_value = float(updated_price) if updated_price not in [None, ''] else None
                except (ValueError, TypeError):
                    price_value = None
                
                size_groups[size_group].append({
                    'row': r,
                    'color': color_str,
                    'clarity': clarity_str,
                    'color_score': color_score,
                    'clarity_score': clarity_score,
                    'total_score': total_quality_score,
                    'price': price_value
                })
        
        # Validate prices within each size group
        validation_row = 2
        total_anomalies = 0
        
        for size_group, stones in size_groups.items():
            # Sort stones by quality score (highest first)
            stones_sorted = sorted(stones, key=lambda x: x['total_score'], reverse=True)
            
            # Calculate expected price ranges based on quality hierarchy
            # Better quality should have HIGHER prices, not lower
            
            # First, find the maximum price among all stones in this group
            all_prices = [s['price'] for s in stones_sorted if s['price'] is not None]
            if not all_prices:
                continue
                
            max_price_in_group = max(all_prices)
            min_price_in_group = min(all_prices)
            avg_price = sum(all_prices) / len(all_prices)
            
            # For each stone, check if its price is appropriate for its quality
            for i, stone in enumerate(stones_sorted):
                expected_max = None
                status = "✓ OK"
                highlight_color = None
                
                if stone['price'] is not None:
                    # Calculate expected price based on quality rank
                    # Higher quality stones should have prices >= lower quality stones
                    quality_rank = i  # 0 = best quality
                    total_stones = len(stones_sorted)
                    
                    # Expected price should decrease as quality decreases
                    # Best quality should have highest price
                    if quality_rank == 0:  # Best quality
                        # Best quality should have price >= 95% of max price
                        expected_max = max_price_in_group
                        if stone['price'] < max_price_in_group * 0.9:
                            status = "⚠️ LOW PRICE - Best quality should have highest price"
                            highlight_color = red_fill
                            total_anomalies += 1
                    else:
                        # Compare with better quality stones
                        better_stones = stones_sorted[:i]
                        better_prices = [s['price'] for s in better_stones if s['price'] is not None]
                        
                        if better_prices:
                            avg_better_price = sum(better_prices) / len(better_prices)
                            # Lower quality should not exceed average of better quality by more than 10%
                            if stone['price'] > avg_better_price * 1.1:
                                status = f"⚠️ HIGH PRICE - Lower quality exceeds better quality avg (${avg_better_price:.0f})"
                                highlight_color = red_fill
                                total_anomalies += 1
                            # Check if price is unreasonably high compared to best quality
                            elif stone['price'] > max_price_in_group * 1.05:
                                status = f"⚠️ PRICE TOO HIGH - Exceeds best quality price"
                                highlight_color = red_fill
                                total_anomalies += 1
                            # Check if price is too low for its quality (optional)
                            elif stone['price'] < avg_price * 0.7 and quality_rank < total_stones // 2:
                                status = f"⚠️ LOW PRICE - Quality position {quality_rank + 1} but price too low"
                                highlight_color = yellow_fill
                    
                    # Highlight the Updated Price cell if anomaly found
                    if highlight_color:
                        cell = ws.cell(row=stone['row'], column=up_col)
                        cell.fill = highlight_color
                        cell.font = Font(color="000000", bold=True)
                
                # Write to validation sheet
                ws_validation.cell(row=validation_row, column=1).value = size_group
                ws_validation.cell(row=validation_row, column=2).value = stone['row']
                ws_validation.cell(row=validation_row, column=3).value = stone['color']
                ws_validation.cell(row=validation_row, column=4).value = stone['clarity']
                ws_validation.cell(row=validation_row, column=5).value = stone['total_score']
                ws_validation.cell(row=validation_row, column=6).value = stone['price']
                ws_validation.cell(row=validation_row, column=8).value = status
                
                if highlight_color == red_fill:
                    ws_validation.cell(row=validation_row, column=8).fill = red_fill
                elif highlight_color == yellow_fill:
                    ws_validation.cell(row=validation_row, column=8).fill = yellow_fill
                
                validation_row += 1
        
        # Add a summary sheet
        if "ValidationSummary" in wb.sheetnames:
            del wb["ValidationSummary"]
        ws_summary = wb.create_sheet("ValidationSummary")
        
        ws_summary.cell(row=1, column=1).value = "PRICE VALIDATION RULES"
        ws_summary.cell(row=1, column=1).font = Font(bold=True, size=14)
        
        ws_summary.cell(row=3, column=1).value = "1. Color Ranking (Best to Worst):"
        ws_summary.cell(row=3, column=2).value = "D > E > F > G > H > I > J"
        
        ws_summary.cell(row=4, column=1).value = "2. Clarity Ranking (Best to Worst):"
        ws_summary.cell(row=4, column=2).value = "IF > VVS1 > VVS2 > VS1 > VS2 > SI1 > SI2"
        
        ws_summary.cell(row=5, column=1).value = "3. Validation Logic:"
        ws_summary.cell(row=5, column=2).value = "Higher quality stones should have HIGHER prices"
        
        ws_summary.cell(row=6, column=1).value = "4. Anomaly Types:"
        ws_summary.cell(row=6, column=2).value = "🔴 RED = Price anomaly (too high for quality)"
        ws_summary.cell(row=7, column=2).value = "🟡 YELLOW = Price too low for quality position"
        
        ws_summary.cell(row=9, column=1).value = f"Total Price Anomalies Found: {total_anomalies}"
        
        if total_anomalies > 0:
            ws_summary.cell(row=9, column=1).fill = red_fill
            ws_summary.cell(row=10, column=1).value = "Check 'PriceValidation' sheet for details"
        
        # Auto-adjust column widths for validation sheet
        for column in ws_validation.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 30)
            ws_validation.column_dimensions[column_letter].width = adjusted_width

    # ── STEP 4: Save & Return ──
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output