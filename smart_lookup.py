# import streamlit as st
# import pandas as pd
# from openpyxl import load_workbook
# import io

# def show_smart_lookup():

#     st.markdown('<div class="logo">⚡ Smart Lookup</div>', unsafe_allow_html=True)

#     main_file = st.file_uploader("Upload Main File", type=["xlsx"])
#     files = st.file_uploader("Upload Team Files", type=["xlsx"], accept_multiple_files=True)

#     if main_file and files:

#         wb = load_workbook(main_file)
#         ws = wb.active

#         df = pd.read_excel(main_file)

#         lot_col = next((c for c in df.columns if "lot" in c.lower()), None)
#         upd_col = next((c for c in df.columns if "updated" in c.lower()), None)

#         lot_idx = list(df.columns).index(lot_col) + 1
#         upd_idx = list(df.columns).index(upd_col) + 1

#         lookup = {}

#         for f in files:
#             temp = pd.read_excel(f)
#             l = next((c for c in temp.columns if "lot" in c.lower()), None)
#             p = next((c for c in temp.columns if "price" in c.lower()), None)

#             if l and p:
#                 for _, row in temp.iterrows():
#                     lookup[row[l]] = row[p]

#         for r in range(2, ws.max_row+1):
#             lot_val = ws.cell(r, lot_idx).value
#             current = ws.cell(r, upd_idx).value

#             if (current is None or str(current).strip()=="") and lot_val in lookup:
#                 ws.cell(r, upd_idx).value = lookup[lot_val]

#         output = io.BytesIO()
#         wb.save(output)
#         output.seek(0)

#         st.download_button("📥 Download Final File", output, "final_output.xlsx")

#     if st.button("⬅ Back"):
#         st.session_state.page = "home"
#         st.rerun()


import streamlit as st
import pandas as pd
from openpyxl import load_workbook
import io

def show_smart_lookup():

    st.markdown('<div class="logo">⚡ Smart Lookup</div>', unsafe_allow_html=True)
    
    # Add option selection
    lookup_option = st.radio(
        "Select Lookup Mode:",
        ["Fill Empty Updated Price", "Update All Prices (Overwrite)"],
        help="Choose how to apply prices from team files"
    )
    
    main_file = st.file_uploader("Upload Main File", type=["xlsx"])
    files = st.file_uploader("Upload Team Files", type=["xlsx"], accept_multiple_files=True)

    if main_file and files:

        wb = load_workbook(main_file)
        ws = wb.active

        df = pd.read_excel(main_file)

        lot_col = next((c for c in df.columns if "lot" in c.lower()), None)
        upd_col = next((c for c in df.columns if "updated" in c.lower()), None)

        lot_idx = list(df.columns).index(lot_col) + 1
        upd_idx = list(df.columns).index(upd_col) + 1

        lookup = {}

        for f in files:
            temp = pd.read_excel(f)
            l = next((c for c in temp.columns if "lot" in c.lower()), None)
            p = next((c for c in temp.columns if "price" in c.lower()), None)

            if l and p:
                for _, row in temp.iterrows():
                    lookup[row[l]] = row[p]

        # Counter for updated cells
        updated_count = 0
        
        for r in range(2, ws.max_row+1):
            lot_val = ws.cell(r, lot_idx).value
            current = ws.cell(r, upd_idx).value
            
            # Check if lot exists in lookup
            if lot_val in lookup:
                if lookup_option == "Fill Empty Updated Price":
                    # Original logic: only fill if empty
                    if current is None or str(current).strip() == "":
                        ws.cell(r, upd_idx).value = lookup[lot_val]
                        updated_count += 1
                else:  # "Update All Prices (Overwrite)"
                    # New logic: update all prices regardless of existing value
                    ws.cell(r, upd_idx).value = lookup[lot_val]
                    updated_count += 1

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        # Show summary
        st.success(f"✅ Updated {updated_count} cells using '{lookup_option}' mode")
        
        st.download_button("📥 Download Final File", output, "final_output.xlsx")

    if st.button("⬅ Back"):
        st.session_state.page = "home"
        st.rerun()